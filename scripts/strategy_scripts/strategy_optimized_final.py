# -*- coding: utf-8 -*-
import time
import akshare as ak
import argparse
import concurrent.futures
import gc
import logging
import os
import traceback
import ssl
from email.mime.base import MIMEBase
from email import encoders
import json
import pickle
from typing import Dict, List, Tuple, Optional

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from psycopg2 import sql
from psycopg2.pool import SimpleConnectionPool
from psycopg2.extras import execute_values

# 机器学习库（可选）
try:
    import xgboost as xgb
    from sklearn.model_selection import train_test_split
    from sklearn.metrics import accuracy_score, roc_auc_score
    from sklearn.preprocessing import StandardScaler
    import joblib
    ML_AVAILABLE = True
except ImportError:
    ML_AVAILABLE = False
    logging.warning("机器学习库未安装，将禁用机器学习功能")

class MLModelLoader:
    """机器学习模型加载器 - 支持自动热重载"""
    _instance = None
    _model = None
    _scaler = None
    _metadata = None
    _last_load_time = 0
    _model_mtime = 0

    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._load_model()
        return cls._instance

    @classmethod
    def _check_model_update(cls):
        """检查模型文件是否有更新"""
        if not os.path.exists(MLConfig.MODEL_PATH):
            return False

        current_mtime = os.path.getmtime(MLConfig.MODEL_PATH)
        if current_mtime > cls._model_mtime:
            cls._model_mtime = current_mtime
            return True
        return False

    @classmethod
    def _load_model(cls):
        """加载或重载模型"""
        if not ML_AVAILABLE:
            return
        try:
            if os.path.exists(MLConfig.MODEL_PATH) and os.path.exists(MLConfig.SCALER_PATH):
                cls._model = joblib.load(MLConfig.MODEL_PATH)
                cls._scaler = joblib.load(MLConfig.SCALER_PATH)
                metadata_path = os.path.join(os.path.dirname(MLConfig.MODEL_PATH), 'metadata.json')
                if os.path.exists(metadata_path):
                    with open(metadata_path, 'r') as f:
                        cls._metadata = json.load(f)
                cls._last_load_time = time.time()
                cls._model_mtime = os.path.getmtime(MLConfig.MODEL_PATH)
                logging.info(f"机器学习模型加载成功: {MLConfig.MODEL_PATH}")
                if cls._metadata:
                    logging.info(f"模型训练时间: {cls._metadata.get('training_date', '未知')}")
                    logging.info(f"验证准确率: {cls._metadata.get('validation_accuracy', 0):.4f}")
            else:
                logging.warning("机器学习模型文件不存在，将禁用机器学习功能")
                cls._model = None
        except Exception as e:
            logging.warning(f"机器学习模型加载失败: {str(e)}，将禁用机器学习功能")
            cls._model = None

    @classmethod
    def reload_if_updated(cls):
        """如果模型有更新则自动重载"""
        if cls._check_model_update():
            logging.info("检测到模型文件更新，正在重载...")
            cls._load_model()

    @classmethod
    def is_available(cls):
        # 每次调用时检查是否需要重载模型
        cls.reload_if_updated()
        return ML_AVAILABLE and cls._model is not None

    @classmethod
    def predict_proba(cls, features):
        """预测上涨概率，返回0-1之间的概率值"""
        if not cls.is_available():
            return 0.5  # 模型不可用时返回中性概率
        try:
            features_scaled = cls._scaler.transform(features)
            return cls._model.predict_proba(features_scaled)[:, 1][0]
        except Exception as e:
            logging.warning(f"机器学习预测失败: {str(e)}")
            return 0.5

# ----------------------------
# 常量定义和配置
# ----------------------------

"""邮件发送配置"""
class EmailConfig:
    SMTP_SERVER = "smtp.qq.com"
    SMTP_PORT = 465
    SMTP_USER = "851448443@qq.com"
    SMTP_PASSWORD = "aofwlgcsobymbfdj"
    RECIPIENTS = ["la9408531@163.com", "1049220782@qq.com","122755347@qq.com"]

"""数据库配置"""
class DbConfig:
    DB_HOST = "localhost"
    DB_PORT = "5431"
    DB_USER = "root"
    DB_PASSWORD = "123629He"
    DB_NAME = "baostock"

class RiskConfig:
    """风险控制配置"""
    MAX_POSITION_PER_STOCK = 0.05  # 单只股票最大仓位5%
    MAX_POSITION_PER_STRATEGY = 0.2  # 单个策略最大仓位20%
    STOP_LOSS_PCT = 0.06  # 止损比例6%
    TAKE_PROFIT_PCT = 0.12  # 止盈比例12%
    TRAILING_STOP_PCT = 0.03  # 跟踪止损3%
    MAX_DRAWDOWN = 0.2  # 最大回撤20%
    TRANSACTION_COST = 0.002  # 交易成本千分之二
    MAX_DAILY_POSITIONS = 10  # 每日最多开仓数量
    ANNUAL_TURNOVER_TARGET = 15  # 年换手率目标15倍
    # 新增风险控制参数
    MAX_PORTFOLIO_DRAWDOWN_LEVEL1 = 0.10  # 总回撤10%时减半仓位
    MAX_PORTFOLIO_DRAWDOWN_LEVEL2 = 0.15  # 总回撤15%时空仓
    CAPITAL_ALLOCATION_MODE = 'dynamic'  # 资金分配模式: dynamic/fixed

class MLConfig:
    """机器学习配置"""
    MODEL_PATH = "./models/xgboost_model.model"
    SCALER_PATH = "./models/scaler.pkl"
    FEATURE_COLS = [
        'ma5', 'ma10', 'ma20', 'ma60', 'ma90', 'ma120',
        'vol_ma5', 'vol_ma10', 'vol_ma20',
        'rsi', 'rsi6', 'rsi12', 'rsi24',
        'macd_hist', 'macd_dif', 'macd_dea',
        'boll_upper', 'boll_lower', 'boll_width',
        'turnover_rate', 'pct_chg', 'high_low_range',
        'close', 'high', 'low', 'volume', 'amount'
    ]
    PROBABILITY_THRESHOLD = 0.65  # 预测上涨概率超过65%才交易

class StrategyConfig:
    """策略参数集中管理"""
    LIMIT_UP_PCT_THRESHOLD = 10  # 涨停阈值(%)
    MA_WINDOWS = {
        'short': 5,  # 短期均线窗口
        'long': 20  # 长期均线窗口
    }
    HOLDING_PERIODS = [1, 3, 5, 10, 20]  # 持有期天数
    VOLUME_THRESHOLDS = {
        'breakout': 2.0,  # 成交量突破标准差倍数
        'resonance': 1.5  # 量价共振放量倍数
    }
    STABILITY_THRESHOLD = 5  # 平稳策略的涨跌幅阈值(%)
    REPORT_COLUMNS = ['策略名称', '策略条件描述', '回测日期范围', '买入次数', '表现',
                      '1 日', '3 日', '5 日', '10 日', '20 日', '标准差', '总体胜率', '盈亏比']
    VOL_WINDOW = 5  # 成交量平均窗口
    RSI_WINDOW = 6  # RSI计算窗口
    MACD_FAST = 14  # MACD快速EMA周期
    MACD_SLOW = 53  # MACD慢速EMA周期
    MACD_SIGNAL = 5  # MACD信号线周期
    MA_SUPPORT_RATIO = 0.98
    MIN_SCORE_THRESHOLD = 6  # 条件评分最低阈值
    RSI_OVERBOUGHT = 70  # RSI超买阈值
    RSI_OVERSOLD = 30  # RSI超卖阈值

    DETAIL_COLUMN_MAPPING = {
        'code': '股票代码',
        'name': '股票名称',
        'trade_date': '交易日期',
        'close': '收盘价',
        'pct_chg': '涨跌幅(%)',
        'open': '开盘价',
        'high': '最高价',
        'low': '最低价',
        'volume': '成交量',
        'amount': '成交额',
        'turnover_rate': '换手率',
        'buy_signal': '买入信号',
        'signal_score': '信号评分',
        'ml_prediction': '机器学习预测概率',
        'return_1d': '1日收益率',
        'return_3d': '3日收益率',
        'return_5d': '5日收益率',
        'return_10d': '10日收益率',
        'return_20d': '20日收益率'
    }

    DESCRIPTIONS = {
        '均线交叉策略': '5日均线上穿20日均线时收阳线，当日成交量大于5日均量50%且连续2日大于5日均量确认资金入场，MACD（6,13,5）柱状线为正且DIF>DEA确认多头趋势,RSI(6)脱离超卖区(>30)但未超买(<70),收盘价在60日线上方，产生买入信号',
        '成交量突破策略': '成交量超过20日均值2倍标准差时产生买入信号',
        '神奇九转策略': '1、满足神奇九转形态(9天内满足低点比较条件,且连续9天收盘价低于4天前的收盘价)2、RSI低于超卖阈值 3、MACD绿柱缩短且仍在0轴下方 4、成交量放大 5、20日均线和60日均线齐头向上，且收盘价接近20日均线获得支撑，收盘确认',
        'N字反包策略': '回调后的收盘价低于前一个上涨波段的高点，且当前收盘价高于该高点时产生买入信号',
        '涨停回调策略': '股票涨停后，后续回调不跌破涨停板的最低价时产生买入信号',
        '连续平稳后涨停策略': '连续10个交易日涨跌幅在5%以内，第11个交易日涨停时买入',
        'MACD月线金叉+20日线策略': '月线MACD金叉且20日线附近放量突破时买入',
        '低位涨停换手率策略': '下跌趋势中止跌后价量齐升时买入',
        '涨停回调量价共振策略': '涨停后缩量且20日均线上行时的特定K线形态买入',
        '孕阳线策略': '上个交易日为跌幅≥7%且实体占比≥70%的大阴线，本交易日开盘价和收盘价均位于前一日开盘价与收盘价之间且收阳线时产生买入信号',
        '多均线共振策略': '1、120、90、60、20日这4条均线都齐头向上，5日均线刚好金叉10日均线2、成交量放大3、MACD（14,53,5）金叉，且量柱由绿翻红4、boll带开口5、RSI三线齐头向上',
        '组合策略': '多策略共振 - 至少2个以下策略同时触发时买入：均线交叉、成交量突破、多均线共振。要求信号综合评分达到阈值，降低假信号风险'
    }


# ----------------------------
# 动态参数管理
# ----------------------------
class DynamicParams:
    """动态参数管理，根据市场环境自动调整策略参数"""
    
    MARKET_REGIMES = ['bull', 'bear', 'volatile', 'structural']
    
    def __init__(self, market_data=None):
        self.market_regime = 'volatile'  # 默认震荡市
        self.current_date = None
        if market_data is not None:
            self.detect_market_regime(market_data)
        
        self._set_params()
    
    def detect_market_regime(self, market_data):
        """检测当前市场环境"""
        if market_data is None or len(market_data) < 60:
            return 'volatile'
            
        recent_data = market_data.tail(60).copy()
        self.current_date = recent_data['trade_date'].max()
        
        # 计算20日收益率
        trend_20d = (recent_data['close'].iloc[-1] / recent_data['close'].iloc[-20] - 1) * 100
        
        # 计算20日波动率
        returns = recent_data['pct_chg'] / 100
        volatility_20d = returns.tail(20).std() * 100
        
        # 计算上涨天数占比
        up_days_pct = (returns > 0).tail(20).mean() * 100
        
        # 市场环境判断
        if trend_20d > 10 and volatility_20d < 2 and up_days_pct > 60:
            self.market_regime = 'bull'
        elif trend_20d < -10 and volatility_20d > 3 and up_days_pct < 40:
            self.market_regime = 'bear'
        elif volatility_20d > 3 and abs(trend_20d) < 5:
            self.market_regime = 'volatile'
        else:
            self.market_regime = 'structural'
            
        logger.info("当前市场环境检测为: %s (20日涨幅: %.2f%%, 波动率: %.2f%%, 上涨天数占比: %.1f%%)" % 
                   (self.market_regime, trend_20d, volatility_20d, up_days_pct))
        return self.market_regime
    
    def _set_params(self):
        """根据市场环境设置参数"""
        if self.market_regime == 'bull':
            self.MIN_SCORE_THRESHOLD = 5
            self.STOP_LOSS_PCT = RiskConfig.STOP_LOSS_PCT * 1.5
            self.TAKE_PROFIT_PCT = RiskConfig.TAKE_PROFIT_PCT * 1.5
            self.MAX_POSITION_PER_STOCK = RiskConfig.MAX_POSITION_PER_STOCK * 1.2
            self.ML_PROBABILITY_THRESHOLD = MLConfig.PROBABILITY_THRESHOLD * 0.9
        elif self.market_regime == 'bear':
            self.MIN_SCORE_THRESHOLD = 8
            self.STOP_LOSS_PCT = RiskConfig.STOP_LOSS_PCT * 0.7
            self.TAKE_PROFIT_PCT = RiskConfig.TAKE_PROFIT_PCT * 0.7
            self.MAX_POSITION_PER_STOCK = RiskConfig.MAX_POSITION_PER_STOCK * 0.5
            self.ML_PROBABILITY_THRESHOLD = MLConfig.PROBABILITY_THRESHOLD * 1.1
        elif self.market_regime == 'volatile':
            self.MIN_SCORE_THRESHOLD = 6
            self.STOP_LOSS_PCT = RiskConfig.STOP_LOSS_PCT
            self.TAKE_PROFIT_PCT = RiskConfig.TAKE_PROFIT_PCT * 0.8
            self.MAX_POSITION_PER_STOCK = RiskConfig.MAX_POSITION_PER_STOCK
            self.ML_PROBABILITY_THRESHOLD = MLConfig.PROBABILITY_THRESHOLD
        else:  # structural
            self.MIN_SCORE_THRESHOLD = 7
            self.STOP_LOSS_PCT = RiskConfig.STOP_LOSS_PCT
            self.TAKE_PROFIT_PCT = RiskConfig.TAKE_PROFIT_PCT
            self.MAX_POSITION_PER_STOCK = RiskConfig.MAX_POSITION_PER_STOCK * 0.8
            self.ML_PROBABILITY_THRESHOLD = MLConfig.PROBABILITY_THRESHOLD * 1.05
            
        logger.info("动态参数已设置: 最低评分阈值=%d, 止损=%.1f%%, 止盈=%.1f%%, 单票仓位上限=%.1f%%, ML概率阈值=%.0f%%" % (
            self.MIN_SCORE_THRESHOLD,
            self.STOP_LOSS_PCT * 100,
            self.TAKE_PROFIT_PCT * 100,
            self.MAX_POSITION_PER_STOCK * 100,
            self.ML_PROBABILITY_THRESHOLD * 100
        ))


# ----------------------------
# 仓位管理模块
# ----------------------------
class PortfolioManager:
    """仓位管理器，负责资金分配和风险控制"""

    def __init__(self, initial_capital=1000000, dynamic_params=None):
        self.initial_capital = initial_capital
        self.current_capital = initial_capital
        self.available_capital = initial_capital
        self.holdings = {}  # {code: {'shares': int, 'cost': float, 'entry_date': date}}
        self.positions = {}  # {code: position_size}
        self.equity_curve = [initial_capital]
        self.drawdown_history = [0]
        self.peak_capital = initial_capital
        self.current_drawdown = 0
        self.dynamic_params = dynamic_params or DynamicParams()

    def calculate_position_size(self, signal_data, available_signals):
        """计算每个信号的仓位大小"""
        position_sizes = {}
        total_signals = len(available_signals)

        if total_signals == 0:
            return position_sizes

        # 根据市场环境调整总仓位
        market_risk_factor = self._calculate_market_risk_factor()

        # 计算可用资金（考虑风险预留）
        total_capital_to_use = self.available_capital * market_risk_factor

        # 每个信号的基准仓位
        base_position = total_capital_to_use / total_signals

        # 限制单只股票最大仓位
        max_position_size = total_capital_to_use * self.dynamic_params.MAX_POSITION_PER_STOCK

        for idx, signal in available_signals.iterrows():
            base_size = min(base_position, max_position_size)

            # 根据信号质量调整仓位
            quality_multiplier = 1.0
            if 'quality_score' in signal:
                quality_multiplier = 0.5 + (signal['quality_score'] / 200)  # 0.5-1.0

            if 'ml_prediction' in signal and pd.notna(signal['ml_prediction']):
                ml_multiplier = 0.5 + signal['ml_prediction']  # 0.5-1.5
                quality_multiplier *= ml_multiplier

            final_position = base_size * quality_multiplier
            final_position = min(final_position, max_position_size)

            position_sizes[signal['code']] = max(final_position, 0)

        return position_sizes

    def _calculate_market_risk_factor(self):
        """根据市场环境和当前回撤计算资金使用率"""
        base_factor = 1.0

        # 根据市场环境调整
        if self.dynamic_params.market_regime == 'bull':
            base_factor = 0.95
        elif self.dynamic_params.market_regime == 'structural':
            base_factor = 0.80
        elif self.dynamic_params.market_regime == 'volatile':
            base_factor = 0.60
        else:  # bear
            base_factor = 0.30

        # 根据回撤调整
        if self.current_drawdown >= RiskConfig.MAX_PORTFOLIO_DRAWDOWN_LEVEL2:
            base_factor = 0.0  # 空仓
        elif self.current_drawdown >= RiskConfig.MAX_PORTFOLIO_DRAWDOWN_LEVEL1:
            base_factor *= 0.5  # 减半

        return base_factor

    def update_equity(self, new_capital):
        """更新资金情况和回撤"""
        self.current_capital = new_capital
        self.equity_curve.append(new_capital)

        # 更新历史最高点和回撤
        if new_capital > self.peak_capital:
            self.peak_capital = new_capital
            self.current_drawdown = 0
        else:
            self.current_drawdown = (self.peak_capital - new_capital) / self.peak_capital

        self.drawdown_history.append(self.current_drawdown)

    def apply_transaction_cost(self, amount):
        """应用交易成本"""
        cost = amount * RiskConfig.TRANSACTION_COST
        return amount - cost


# ----------------------------
# 绩效指标计算模块
# ----------------------------
class PerformanceMetrics:
    """策略回测绩效指标计算"""

    @staticmethod
    def calculate_all_metrics(equity_curve, returns_series, risk_free_rate=0.03):
        """计算所有绩效指标"""
        equity_array = np.array(equity_curve)
        returns_array = np.array(returns_series)

        metrics = {}

        # 基础收益指标
        metrics['total_return'] = (equity_array[-1] / equity_array[0] - 1) * 100
        metrics['annualized_return'] = PerformanceMetrics._annualized_return(
            equity_array, period_days=len(returns_array)
        )

        # 风险指标
        metrics['max_drawdown'] = PerformanceMetrics._max_drawdown(equity_array) * 100
        metrics['volatility'] = np.std(returns_array) * np.sqrt(252) * 100

        # 风险调整收益
        excess_returns = returns_array - risk_free_rate / 252
        if len(excess_returns) > 0 and np.std(excess_returns) > 0:
            metrics['sharpe_ratio'] = np.mean(excess_returns) / np.std(excess_returns) * np.sqrt(252)
        else:
            metrics['sharpe_ratio'] = 0

        # 索提诺比率（只考虑下行风险）
        downside_returns = np.array([r for r in returns_array if r < 0])
        if len(downside_returns) > 0 and np.std(downside_returns) > 0:
            metrics['sortino_ratio'] = np.mean(returns_array) / np.std(downside_returns) * np.sqrt(252)
        else:
            metrics['sortino_ratio'] = 0

        # 胜率和盈亏比
        positive_returns = returns_array[returns_array > 0]
        negative_returns = returns_array[returns_array < 0]

        if len(returns_array) > 0:
            metrics['win_rate'] = len(positive_returns) / len(returns_array) * 100
        else:
            metrics['win_rate'] = 0

        if len(negative_returns) > 0:
            metrics['profit_loss_ratio'] = abs(np.mean(positive_returns)) / abs(np.mean(negative_returns))
        else:
            metrics['profit_loss_ratio'] = float('inf') if len(positive_returns) > 0 else 0

        # 卡尔玛比率
        if metrics['max_drawdown'] != 0:
            metrics['calmar_ratio'] = abs(metrics['annualized_return']) / abs(metrics['max_drawdown'])
        else:
            metrics['calmar_ratio'] = 0

        return metrics

    @staticmethod
    def _annualized_return(equity_array, period_days):
        """计算年化收益率"""
        if len(equity_array) < 2 or equity_array[-1] == 0:
            return 0
        total_return = equity_array[-1] / equity_array[0] - 1
        years = period_days / 252
        if years > 0:
            return ((1 + total_return) ** (1 / years) - 1) * 100
        return 0

    @staticmethod
    def _max_drawdown(equity_array):
        """计算最大回撤"""
        if len(equity_array) < 2:
            return 0

        peak = np.maximum.accumulate(equity_array)
        drawdown = (peak - equity_array) / peak
        return np.max(drawdown)

    @staticmethod
    def print_metrics(metrics, strategy_name="策略"):
        """打印绩效指标"""
        print("\n" + "=" * 50)
        print(f"{strategy_name} 绩效指标")
        print("=" * 50)
        print(f"总收益率:     {metrics.get('total_return', 0):.2f}%")
        print(f"年化收益率:   {metrics.get('annualized_return', 0):.2f}%")
        print(f"最大回撤:     {metrics.get('max_drawdown', 0):.2f}%")
        print(f"年化波动率:   {metrics.get('volatility', 0):.2f}%")
        print(f"夏普比率:     {metrics.get('sharpe_ratio', 0):.2f}")
        print(f"索提诺比率:   {metrics.get('sortino_ratio', 0):.2f}")
        print(f"卡尔玛比率:   {metrics.get('calmar_ratio', 0):.2f}")
        print(f"胜率:         {metrics.get('win_rate', 0):.2f}%")
        print(f"盈亏比:       {metrics.get('profit_loss_ratio', 0):.2f}")
        print("=" * 50 + "\n")


# ----------------------------
# 策略注册机制
# ----------------------------
class StrategyRegistry:
    """策略注册器，集中管理所有交易策略"""
    _strategies = {}

    @classmethod
    def register(cls, name, description=""):
        def decorator(func):
            cls._strategies[name] = {
                'function': func,
                'description': description or StrategyConfig.DESCRIPTIONS.get(name, "")
            }
            return func
        return decorator

    @classmethod
    def get_strategy(cls, name):
        return cls._strategies.get(name, {}).get('function')

    @classmethod
    def get_description(cls, name):
        return cls._strategies.get(name, {}).get('description', "")

    @classmethod
    def list_strategies(cls):
        return list(cls._strategies.keys())


# ----------------------------
# 日志配置
# ----------------------------
def setup_logger():
    """配置日志系统"""
    logger = logging.getLogger('strategy_backtest')
    logger.setLevel(logging.DEBUG)

    log_dir = "./result"
    os.makedirs(log_dir, exist_ok=True)

    file_handler = logging.FileHandler(
        os.path.join(log_dir, "backtest.log"),
        encoding='utf-8'
    )
    file_handler.setLevel(logging.DEBUG)

    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)

    formatter = logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)

    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    return logger


logger = setup_logger()


# ----------------------------
# 市场环境过滤器
# ----------------------------
class MarketFilter:
    """市场环境过滤器 - 过滤不适合交易的日子"""

    def __init__(self, return_data=None):
        self.return_data = return_data

    def get_market_sentiment(self, date):
        """获取市场整体情绪（0-1，越高表示市场越暖）"""
        if self.return_data is None:
            return 0.5

        date_data = self.return_data[self.return_data['date'] == date]
        if date_data.empty:
            return 0.5

        up_count = len(date_data[date_data['pct_chg'] > 0])
        total_count = len(date_data)
        sentiment = up_count / total_count if total_count > 0 else 0.5

        return sentiment

    def get_market_index_trend(self, date, window=20):
        """获取市场指数趋势"""
        if self.return_data is None:
            return 0

        date_data = self.return_data[self.return_data['date'] <= date].tail(window)
        if len(date_data) < window:
            return 0

        avg_returns = date_data.groupby('date')['pct_chg'].mean()
        if len(avg_returns) < 2:
            return 0

        recent_trend = (avg_returns.iloc[-1] - avg_returns.iloc[0])
        return recent_trend

    def should_trade(self, date, min_sentiment=0.5):
        """判断当日是否适合交易"""
        sentiment = self.get_market_sentiment(date)
        trend = self.get_market_index_trend(date)

        if sentiment >= min_sentiment:
            return True

        if sentiment >= min_sentiment - 0.15 and trend > 0.5:
            return True

        logger.info(f"日期 {date} 市场情绪: {sentiment:.2%}, 趋势: {trend:.2f}% - 不交易")
        return False

    def filter_by_market_condition(self, buy_signals, min_sentiment=0.5):
        """根据市场环境过滤买入信号"""
        if buy_signals.empty:
            return buy_signals

        original_count = len(buy_signals)
        valid_dates = []

        for date in buy_signals['date'].unique():
            if self.should_trade(date, min_sentiment):
                valid_dates.append(date)

        filtered = buy_signals[buy_signals['date'].isin(valid_dates)].copy()
        logger.info(f"市场环境过滤: 从 {original_count} 个信号减少到 {len(filtered)} 个")

        return filtered


# ----------------------------
# 信号质量评分系统
# ----------------------------
class SignalQualityScorer:
    """信号质量评分系统 - 对单一买入信号进行质量打分（0-100）"""

    def __init__(self, return_data=None):
        self.return_data = return_data
        # Ensure code column exists
        if return_data is not None and 'code' not in return_data.columns:
            # Try to find code column by other names
            if 'stock_code' in return_data.columns:
                return_data.rename(columns={'stock_code': 'code'}, inplace=True)
            elif 'ts_code' in return_data.columns:
                return_data.rename(columns={'ts_code': 'code'}, inplace=True)

    def calculate_signal_quality(self, buy_signal_row, stock_history=None, enable_ml=True):
        """计算信号质量分数"""
        if stock_history is None or stock_history.empty:
            return 50

        latest = stock_history.iloc[-1]
        score = 0

        # 0. 机器学习预测分数（25分）
        if enable_ml and MLModelLoader.is_available():
            try:
                # 提取特征
                feature_values = []
                for col in MLConfig.FEATURE_COLS:
                    if col in latest:
                        val = latest[col]
                        # 处理异常值
                        if pd.isna(val) or np.isinf(val) or np.isneginf(val):
                            val = 0
                        feature_values.append(val)
                    else:
                        feature_values.append(0)

                # 转换为模型输入格式
                features = np.array(feature_values).reshape(1, -1)
                prob = MLModelLoader.predict_proba(features)

                # 根据预测概率给分
                if prob >= 0.75:
                    score += 25
                elif prob >= 0.70:
                    score += 20
                elif prob >= 0.65:
                    score += 15
                elif prob >= 0.60:
                    score += 10
                elif prob >= 0.55:
                    score += 5

            except Exception as e:
                logging.debug(f"ML预测失败: {str(e)}")

        # 1. 成交量确认（30分）
        try:
            vol_ma5 = latest.get('vol_ma5', 0)
            current_vol = latest.get('volume', 0)

            if vol_ma5 > 0:
                vol_ratio = current_vol / vol_ma5
                if vol_ratio >= 2.0:
                    score += 30
                elif vol_ratio >= 1.5:
                    score += 25
                elif vol_ratio >= 1.2:
                    score += 15
                elif vol_ratio >= 1.0:
                    score += 5
        except Exception:
            pass

        # 2. 均线排列确认（25分）
        try:
            ma5 = latest.get('ma5', 0)
            ma10 = latest.get('ma10', 0)
            ma20 = latest.get('ma20', 0)
            ma60 = latest.get('ma60', 0)

            if ma5 > 0 and ma10 > 0 and ma20 > 0:
                if ma5 > ma10 > ma20:
                    score += 20
                    if ma20 > ma60:
                        score += 5
                elif ma5 > ma10:
                    score += 10
        except Exception:
            pass

        # 3. RSI区间确认（20分）
        try:
            rsi6 = latest.get('rsi6', 0)
            rsi12 = latest.get('rsi12', 0)

            if 30 < rsi6 < 70:
                score += 15
                if 40 < rsi6 < 60:
                    score += 5

            if 30 < rsi12 < 70:
                score += 5
        except Exception:
            pass

        # 4. 位置确认（15分）
        try:
            close = latest.get('close', 0)
            ma5 = latest.get('ma5', 0)
            ma10 = latest.get('ma10', 0)
            high_60d = stock_history['high'].tail(60).max()

            if close > 0 and ma5 > 0:
                if 0.98 < close / ma5 < 1.02:
                    score += 10
                elif close > high_60d:
                    score += 15
                elif 0.95 < close / ma10 < 1.05:
                    score += 8
        except Exception:
            pass

        # 5. 波动率确认（10分）
        try:
            volatility = stock_history['close'].tail(20).std() / latest.get('ma20', 1)
            if 0.01 < volatility < 0.035:
                score += 10
            elif 0.035 <= volatility < 0.05:
                score += 5
        except Exception:
            pass

        return min(score, 100)

    def filter_by_quality(self, buy_signals, min_score=65, enable_ml=True):
        """只保留高分信号"""
        if buy_signals.empty:
            return buy_signals

        quality_scores = []
        original_count = len(buy_signals)

        logger.info("开始计算信号质量评分，这可能需要一些时间...")

        for idx, row in buy_signals.iterrows():
            try:
                stock_history = self.return_data[self.return_data['code'] == row['code']]
                signal_date = row['date']
                history = stock_history[stock_history['date'] < signal_date].tail(120)
                if not history.empty:
                    score = self.calculate_signal_quality(row, history, enable_ml=enable_ml)
                else:
                    score = 50
                quality_scores.append(score)
            except Exception as e:
                logger.warning(f"计算信号质量失败 (股票: {row['code']}): {e}")
                quality_scores.append(50)

        buy_signals['quality_score'] = quality_scores
        filtered = buy_signals[buy_signals['quality_score'] >= min_score].copy()

        logger.info(f"信号质量过滤: 从 {original_count} 个信号减少到 {len(filtered)} 个 (最低分数: {min_score})")

        return filtered

    def add_ml_probability(self, buy_signals):
        """为信号添加机器学习上涨概率"""
        if not MLModelLoader.is_available() or buy_signals.empty:
            buy_signals['ml_probability'] = 0.5
            return buy_signals

        ml_probs = []
        for idx, row in buy_signals.iterrows():
            try:
                stock_history = self.return_data[self.return_data['code'] == row['code']]
                signal_date = row['date']
                history = stock_history[stock_history['date'] < signal_date].tail(1)
                if not history.empty:
                    latest = history.iloc[-1]
                    feature_values = []
                    for col in MLConfig.FEATURE_COLS:
                        if col in latest:
                            val = latest[col]
                            if pd.isna(val) or np.isinf(val) or np.isneginf(val):
                                val = 0
                            feature_values.append(val)
                        else:
                            feature_values.append(0)
                    features = np.array(feature_values).reshape(1, -1)
                    prob = MLModelLoader.predict_proba(features)
                    ml_probs.append(prob)
                else:
                    ml_probs.append(0.5)
            except Exception as e:
                logger.debug(f"ML概率计算失败 (股票: {row['code']}): {e}")
                ml_probs.append(0.5)

        buy_signals['ml_probability'] = ml_probs
        return buy_signals

    def filter_by_ml(self, buy_signals, min_probability=None):
        """根据机器学习预测概率过滤信号"""
        if not MLModelLoader.is_available() or buy_signals.empty:
            return buy_signals

        if min_probability is None:
            min_probability = MLConfig.PROBABILITY_THRESHOLD

        if 'ml_probability' not in buy_signals.columns:
            buy_signals = self.add_ml_probability(buy_signals)

        original_count = len(buy_signals)
        filtered = buy_signals[buy_signals['ml_probability'] >= min_probability].copy()

        logger.info(f"机器学习过滤: 从 {original_count} 个信号减少到 {len(filtered)} 个 (最低概率: {min_probability:.2f})")

        return filtered


# ----------------------------
# 止损止盈机制
# ----------------------------
def calculate_return_with_stop_loss(returns_series, stop_loss_pct=0.06, take_profit_pct=0.12,
                                    trailing_stop_pct=0.03):
    """考虑止损止盈的收益率计算（使用跟踪止损）"""
    if not hasattr(returns_series, '__iter__') or len(returns_series) == 0:
        return pd.Series(returns_series)

    adjusted_returns = []
    max_return_seen = 0

    for ret in returns_series:
        if ret > max_return_seen:
            max_return_seen = ret

        if ret <= -stop_loss_pct:
            adjusted_returns.append(-stop_loss_pct)
            max_return_seen = 0
        elif ret >= take_profit_pct:
            adjusted_returns.append(take_profit_pct)
            max_return_seen = 0
        elif max_return_seen - ret >= trailing_stop_pct and max_return_seen > trailing_stop_pct:
            adjusted_returns.append(ret)
            max_return_seen = ret
        else:
            adjusted_returns.append(ret)

    return pd.Series(adjusted_returns)


# ----------------------------
# 交易日判定
# ----------------------------
def is_trading_day():
    today = datetime.now().strftime('%Y-%m-%d')
    try:
        stock_trade_calendar_df = ak.tool_trade_date_hist_sina()
        stock_trade_calendar_df['trade_date'] = pd.to_datetime(stock_trade_calendar_df['trade_date'])
        today_date = pd.to_datetime(today)
        is_trading = today_date in stock_trade_calendar_df['trade_date'].values
        logger.info("今天是交易日: %s" % str(is_trading))
        return is_trading
    except Exception as e:
        logger.error("获取交易日历出错: %s" % str(e))
        return False

# 新增独立函数用于策略应用（增强版，含止损止盈、交易成本、风险控制）
def apply_strategy_independent(stock_data, strategy_func, strategy_name="策略",
                               dynamic_params=None, ml_filter=None, enable_stop_loss=True,
                               enable_transaction_cost=True):
    """独立于数据库连接的应用策略函数（含止损止盈、交易成本、风险控制机制）"""
    if stock_data is None or stock_data.empty:
        logger.warning("%s: 没有数据用于策略评估" % strategy_name)
        return None, None

    try:
        logger.info("开始应用策略: %s" % strategy_name)
        strategy_data = strategy_func(stock_data.copy(), dynamic_params)

        if 'buy_signal' not in strategy_data.columns:
            logger.warning("%s: 策略未生成buy_signal列" % strategy_name)
            return strategy_data, None

        buy_signals = strategy_data[strategy_data['buy_signal'] == 1].copy()
        logger.info("%s: 生成 %d 个买入信号" % (strategy_name, len(buy_signals)))

        if buy_signals.empty:
            logger.warning("%s: 没有有效的买入信号" % strategy_name)
            return strategy_data, None

        # 信号质量过滤（包含机器学习评分）
        try:
            scorer = SignalQualityScorer(stock_data)

            original_count = len(buy_signals)
            logger.info(f"原始信号数量: {original_count}")

            # ML过滤（只有当启用时才应用）
            if ml_filter and MLModelLoader.is_available():
                buy_signals = scorer.add_ml_probability(buy_signals)
                buy_signals = scorer.filter_by_ml(buy_signals)
                logger.info(f"ML过滤后剩余信号: {len(buy_signals)}")

            # 质量评分过滤
            min_score = 60  # 使用相同的质量阈值进行公平比较
            buy_signals = scorer.filter_by_quality(buy_signals, min_score=min_score, enable_ml=ml_filter)
            logger.info(f"质量过滤后剩余信号: {len(buy_signals)}")

            if buy_signals.empty:
                logger.warning("%s: 经过过滤后没有剩余信号" % strategy_name)
                return strategy_data, None
        except Exception as e:
            logger.warning(f"信号质量过滤出错: {e}，将使用原始信号")

        results = {
            '策略名称': strategy_name,
            '买入次数': len(buy_signals),
            '策略描述': StrategyConfig.DESCRIPTIONS.get(strategy_name, "")
        }

        # 应用止损止盈参数
        stop_loss_pct = dynamic_params.STOP_LOSS_PCT if dynamic_params else RiskConfig.STOP_LOSS_PCT
        take_profit_pct = dynamic_params.TAKE_PROFIT_PCT if dynamic_params else RiskConfig.TAKE_PROFIT_PCT
        trailing_stop_pct = RiskConfig.TRAILING_STOP_PCT
        transaction_cost = RiskConfig.TRANSACTION_COST if enable_transaction_cost else 0

        logger.info(f"应用风险控制参数: 止损={stop_loss_pct*100:.1f}%, 止盈={take_profit_pct*100:.1f}%, 跟踪止损={trailing_stop_pct*100:.1f}%, 交易成本={transaction_cost*100:.2f}%")

        # 收集所有持仓期的收益用于绩效评估
        all_returns = []

        for period in StrategyConfig.HOLDING_PERIODS:
            return_col = 'return_%dd' % period
            if return_col not in buy_signals.columns:
                logger.warning("%s: 缺少%s列，跳过该持有期" % (strategy_name, return_col))
                continue

            if enable_stop_loss:
                adjusted_returns = calculate_return_with_stop_loss(
                    buy_signals[return_col],
                    stop_loss_pct=stop_loss_pct,
                    take_profit_pct=take_profit_pct,
                    trailing_stop_pct=trailing_stop_pct
                )
            else:
                adjusted_returns = buy_signals[return_col]

            # 应用交易成本
            if enable_transaction_cost:
                adjusted_returns = adjusted_returns - (2 * transaction_cost)  # 买卖双向成本
                logger.info(f"已应用交易成本 ({transaction_cost*100:.2f}%双向)")

            valid_returns = adjusted_returns.replace([np.inf, -np.inf], np.nan).dropna()

            if valid_returns.empty:
                results['%d日上涨概率(%%)' % period] = 0
                results['%d日平均收益率(%%)' % period] = 0
                results['%d日标准差(%%)' % period] = 0
                continue

            # 收集用于绩效统计
            all_returns.extend(valid_returns.tolist())

            pos_returns = valid_returns[valid_returns > 0]
            win_prob = len(pos_returns) / len(valid_returns) * 100

            try:
                avg_return = valid_returns.mean() * 100
                std_dev = valid_returns.std() * 100
            except Exception as e:
                logger.warning("计算收益率统计出错: %s" % str(e))
                avg_return = 0
                std_dev = 0

            results['%d日上涨概率(%%)' % period] = win_prob
            results['%d日平均收益率(%%)' % period] = avg_return
            results['%d日标准差(%%)' % period] = std_dev

        # 添加全周期绩效指标
        if all_returns:
            returns_series = pd.Series(all_returns)
            results['总体胜率(%%)'] = (returns_series > 0).sum() / len(returns_series) * 100
            results['总体平均收益率(%%)'] = returns_series.mean() * 100
            results['总体标准差(%%)'] = returns_series.std() * 100
            results['盈亏比'] = abs(returns_series[returns_series > 0].mean()) / abs(returns_series[returns_series < 0].mean()) if (returns_series < 0).sum() > 0 else 0

        return buy_signals, results

    except Exception as e:
        logger.error("应用策略 %s 时出错: %s" % (strategy_name, str(e)))
        logger.error(traceback.format_exc())
        return None, None


# ----------------------------
# 回测主类
# ----------------------------
class StockStrategyBacktest:
    def __init__(self, db_params, pool_size=5):
        self.db_params = db_params
        self.pool_size = pool_size
        self.connection_pool = None
        self.strategy_results = {}
        self.strategy_buy_details = {}
        self.stock_data = None
        self.return_data = None
        self.report_file_path = None
        self.dynamic_params = DynamicParams()
        self.risk_manager = None
        self.portfolio_manager = PortfolioManager(dynamic_params=self.dynamic_params)
        self.performance_metrics = {}

        # 初始化数据库表
        self._init_db_tables()

    def _init_db_tables(self):
        """初始化数据库结果表"""
        if not self.create_connection_pool():
            return
            
        conn = self.get_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute("""
                CREATE TABLE IF NOT EXISTS strategy_backtest_results (
                    id SERIAL PRIMARY KEY,
                    strategy_name VARCHAR(100) NOT NULL,
                    start_date DATE NOT NULL,
                    end_date DATE NOT NULL,
                    buy_count INTEGER NOT NULL,
                    win_rate_1d NUMERIC(5,2),
                    win_rate_3d NUMERIC(5,2),
                    win_rate_5d NUMERIC(5,2),
                    win_rate_10d NUMERIC(5,2),
                    win_rate_20d NUMERIC(5,2),
                    avg_return_1d NUMERIC(8,4),
                    avg_return_3d NUMERIC(8,4),
                    avg_return_5d NUMERIC(8,4),
                    avg_return_10d NUMERIC(8,4),
                    avg_return_20d NUMERIC(8,4),
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(strategy_name, start_date, end_date)
                )
                """)
                
                cursor.execute("""
                CREATE TABLE IF NOT EXISTS strategy_signals (
                    id SERIAL PRIMARY KEY,
                    strategy_name VARCHAR(100) NOT NULL,
                    code VARCHAR(20) NOT NULL,
                    name VARCHAR(100),
                    trade_date DATE NOT NULL,
                    close_price NUMERIC(10,2),
                    signal_score INTEGER,
                    ml_prediction NUMERIC(5,4),
                    return_1d NUMERIC(8,4),
                    return_3d NUMERIC(8,4),
                    return_5d NUMERIC(8,4),
                    return_10d NUMERIC(8,4),
                    return_20d NUMERIC(8,4),
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(strategy_name, code, trade_date)
                )
                """)
                
                conn.commit()
                logger.info("数据库表初始化完成")
        except Exception as e:
            logger.error("初始化数据库表失败: %s" % str(e))
            conn.rollback()
        finally:
            self.release_connection(conn)
    
    def save_results_to_db(self, start_date, end_date):
        """保存回测结果到数据库"""
        if not self.strategy_results:
            return
            
        if not self.create_connection_pool():
            return
            
        conn = self.get_connection()
        try:
            with conn.cursor() as cursor:
                for strategy_name, results in self.strategy_results.items():
                    data = (
                        strategy_name,
                        start_date,
                        end_date,
                        results['买入次数'],
                        results.get('1日上涨概率(%)', 0),
                        results.get('3日上涨概率(%)', 0),
                        results.get('5日上涨概率(%)', 0),
                        results.get('10日上涨概率(%)', 0),
                        results.get('20日上涨概率(%)', 0),
                        results.get('1日平均收益率(%)', 0),
                        results.get('3日平均收益率(%)', 0),
                        results.get('5日平均收益率(%)', 0),
                        results.get('10日平均收益率(%)', 0),
                        results.get('20日平均收益率(%)', 0)
                    )
                    
                    cursor.execute("""
                    INSERT INTO strategy_backtest_results (
                        strategy_name, start_date, end_date, buy_count,
                        win_rate_1d, win_rate_3d, win_rate_5d, win_rate_10d, win_rate_20d,
                        avg_return_1d, avg_return_3d, avg_return_5d, avg_return_10d, avg_avg_return_20d
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (strategy_name, start_date, end_date) DO UPDATE SET
                        buy_count = EXCLUDED.buy_count,
                        win_rate_1d = EXCLUDED.win_rate_1d,
                        win_rate_3d = EXCLUDED.win_rate_3d,
                        win_rate_5d = EXCLUDED.win_rate_5d,
                        win_rate_10d = EXCLUDED.win_rate_10d,
                        win_rate_20d = EXCLUDED.win_rate_20d,
                        avg_return_1d = EXCLUDED.avg_return_1d,
                        avg_return_3d = EXCLUDED.avg_return_3d,
                        avg_return_5d = EXCLUDED.avg_return_5d,
                        avg_return_10d = EXCLUDED.avg_return_10d,
                        avg_return_20d = EXCLUDED.avg_return_20d,
                        created_at = CURRENT_TIMESTAMP
                    """, data)
                    
            conn.commit()
            logger.info("回测结果已保存到数据库")
        except Exception as e:
            logger.error("保存结果到数据库失败: %s" % str(e))
            conn.rollback()
        finally:
            self.release_connection(conn)

    def create_connection_pool(self):
        """创建数据库连接池"""
        if self.connection_pool is None:
            try:
                self.connection_pool = SimpleConnectionPool(
                    minconn=1,
                    maxconn=self.pool_size,
                    **self.db_params
                )
                logger.info("成功创建数据库连接池 (大小: %d)" % self.pool_size)
                return True
            except Exception as e:
                logger.error("创建数据库连接池失败: %s" % str(e))
                return False
        return True

    def get_connection(self):
        """从连接池获取数据库连接"""
        if self.connection_pool is None:
            self.create_connection_pool()
        return self.connection_pool.getconn()

    def release_connection(self, conn):
        """释放数据库连接回连接池"""
        if conn:
            self.connection_pool.putconn(conn)

    def close_pool(self):
        """关闭数据库连接池"""
        if self.connection_pool:
            self.connection_pool.closeall()
            logger.info("数据库连接池已关闭")

    def fetch_stock_data(self, start_date, end_date, chunk_size=10000):
        """从数据库获取股票数据"""
        if self.stock_data is not None:
            return self.stock_data

        if not self.create_connection_pool():
            return None

        conn = self.get_connection()
        try:
            with conn.cursor() as cursor:
                query = sql.SQL("""
                                SELECT code,name,
                                       date as trade_date, open, high, low, close, pct_chg, volume, amount, turn as turnover_rate
                                FROM baostock_daily_history
                                WHERE date BETWEEN %s
                                  AND %s
                                ORDER BY code, trade_date
                                """)
                cursor.execute(query, (start_date, end_date))

                chunks = []
                while True:
                    data = cursor.fetchmany(chunk_size)
                    if not data:
                        break
                    columns = [desc[0] for desc in cursor.description]

                    dtype_mapping = {
                        'code': 'category',
                        'name': 'category',
                        'trade_date': 'datetime64[ns]',
                        'open': 'float32',
                        'high': 'float32',
                        'low': 'float32',
                        'close': 'float32',
                        'pct_chg': 'float32',
                        'volume': 'int64',
                        'amount': 'float32',
                        'turnover_rate': 'float32'
                    }

                    df_chunk = pd.DataFrame(data, columns=columns)
                    for col, dtype in dtype_mapping.items():
                        if col in df_chunk.columns:
                            if col == 'volume':
                                # Handle None values for volume
                                df_chunk[col] = pd.to_numeric(df_chunk[col], errors='coerce').fillna(0).astype('int64')
                            else:
                                df_chunk[col] = df_chunk[col].astype(dtype)

                    chunks.append(df_chunk)
                    del df_chunk

                if chunks:
                    df = pd.concat(chunks, ignore_index=True)
                    df['trade_date'] = pd.to_datetime(df['trade_date'])
                    logger.info("成功获取 %d 条股票数据 (%d 只股票)" % (len(df), len(df['code'].unique())))
                    self.stock_data = df
                    return df
                else:
                    logger.warning("未获取到任何股票数据")
                    return None
        except Exception as e:
            logger.error("获取股票数据失败: %s" % str(e))
            return None
        finally:
            self.release_connection(conn)

    def calculate_returns(self, stock_data=None, holding_periods=StrategyConfig.HOLDING_PERIODS):
        """计算不同持有期的收益率"""
        if self.return_data is not None:
            return self.return_data

        df = stock_data if stock_data is not None else self.stock_data
        if df is None or df.empty:
            logger.warning("没有数据用于计算收益率")
            return None

        df = df.sort_values(['code', 'trade_date']).reset_index(drop=True)
        df['close'] = pd.to_numeric(df['close'], errors='coerce')

        logger.info("计算持有期收益率...")
        # Calculate returns using explicit iteration to preserve columns
        result_dfs = []
        for code in df['code'].unique():
            group = df[df['code'] == code].copy()
            for period in holding_periods:
                group[f'return_{period}d'] = group['close'].pct_change(periods=period).shift(-period)
            result_dfs.append(group)
        df = pd.concat(result_dfs, ignore_index=True)

        logger.info("计算技术指标...")
        # Calculate technical indicators
        result_dfs = []
        for code in df['code'].unique():
            group = df[df['code'] == code].copy()
            group = self._calculate_technical_indicators(group)
            result_dfs.append(group)
        df = pd.concat(result_dfs, ignore_index=True)

        cols_to_drop = ['ma5', 'ma20', 'vol_mean', 'vol_std', 'lower_than_4days_ago', 'count']
        existing_cols = [col for col in cols_to_drop if col in df.columns]
        if existing_cols:
            df = df.drop(columns=existing_cols)

        logger.info("收益率和技术指标计算完成，涵盖 %d 只股票" % len(df['code'].unique()))
        self.return_data = df
        return df
        
    def _calculate_technical_indicators(self, group):
        """计算技术指标"""
        group = group.sort_values('trade_date')
        
        for window in [5, 10, 20, 60, 90, 120]:
            group['ma%d' % window] = group['close'].rolling(window=window).mean()
            
        for window in [5, 10, 20]:
            group['vol_ma%d' % window] = group['volume'].rolling(window=window).mean()
            
        delta = group['close'].diff()
        gain = delta.where(delta > 0, 0)
        loss = -delta.where(delta < 0, 0)
        
        for window in [6, 12, 24]:
            avg_gain = gain.rolling(window=window).mean()
            avg_loss = loss.rolling(window=window).mean()
            rs = avg_gain / avg_loss.replace(0, 0.001)
            group['rsi%d' % window] = 100 - (100 / (1 + rs))
            
        ema_fast = group['close'].ewm(span=StrategyConfig.MACD_FAST, adjust=False).mean()
        ema_slow = group['close'].ewm(span=StrategyConfig.MACD_SLOW, adjust=False).mean()
        group['macd_dif'] = ema_fast - ema_slow
        group['macd_dea'] = group['macd_dif'].ewm(span=StrategyConfig.MACD_SIGNAL, adjust=False).mean()
        group['macd_hist'] = 2 * (group['macd_dif'] - group['macd_dea'])
        
        ma20 = group['ma20']
        std20 = group['close'].rolling(window=20).std()
        group['boll_upper'] = ma20 + 2 * std20
        group['boll_lower'] = ma20 - 2 * std20
        group['boll_width'] = (group['boll_upper'] - group['boll_lower']) / ma20
        
        group['high_low_range'] = (group['high'] - group['low']) / group['open']
        group['rsi'] = group['rsi6']
        
        return group

    def apply_strategy(self, stock_data, strategy_func, strategy_name="策略", ml_filter=True, enable_stop_loss=True, enable_transaction_cost=True):
        """应用交易策略并评估结果"""
        buy_signals, results = apply_strategy_independent(
            stock_data,
            strategy_func,
            strategy_name,
            self.dynamic_params,
            ml_filter=ml_filter,
            enable_stop_loss=enable_stop_loss,
            enable_transaction_cost=enable_transaction_cost
        )

        if results is not None:
            self.strategy_results[strategy_name] = results

        if buy_signals is not None:
            self.strategy_buy_details[strategy_name] = buy_signals

        return stock_data, buy_signals

    def generate_report(self, output_format='excel', output_file='strategy_report', start_date='', end_date=''):
        """生成回测报告"""
        if not self.strategy_results:
            logger.warning("没有策略回测结果可用于生成报告")
            return False

        report_dir = "../result"
        os.makedirs(report_dir, exist_ok=True)

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        header_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        data_alignment = Alignment(horizontal='center', vertical='center')

        master_wb = Workbook()
        master_wb.remove(master_wb.active)

        for strategy_name, results in self.strategy_results.items():
            safe_name = strategy_name[:31]
            ws = master_wb.create_sheet(title="%s表现" % safe_name)

            for col_idx, col_name in enumerate(StrategyConfig.REPORT_COLUMNS, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            date_range = '%s 至 %s' % (start_date, end_date)

            row_data_1 = [
                strategy_name,
                results.get('策略描述', ""),
                date_range,
                results['买入次数'],
                '上涨概率',
                results.get('1日上涨概率(%)', 0),
                results.get('3日上涨概率(%)', 0),
                results.get('5日上涨概率(%)', 0),
                results.get('10日上涨概率(%)', 0),
                results.get('20日上涨概率(%)', 0),
                '详情数据更新中',
                results.get('总体胜率(%)', 0),
                results.get('盈亏比', 0)
            ]

            for col_idx, value in enumerate(row_data_1, 1):
                cell = ws.cell(row=2, column=col_idx, value=value)
                cell.alignment = data_alignment
                cell.border = thin_border

            ws.cell(row=3, column=5, value='平均收益率').alignment = data_alignment
            ws.cell(row=3, column=11, value='波动率(5日)').alignment = data_alignment

            for idx, period in enumerate(StrategyConfig.HOLDING_PERIODS, 6):
                ws.cell(row=3, column=idx, value=results.get('%d日平均收益率(%%)' % period, 0))

            ws.cell(row=3, column=12, value=results.get('%d日平均收益率(%%)' % StrategyConfig.HOLDING_PERIODS[2], 0))

            for row in ws.iter_rows(min_row=2, max_row=3):
                for cell in row:
                    cell.alignment = data_alignment
                    cell.alignment = Alignment(wrap_text=True)
                    cell.border = thin_border

            for col_idx in range(1, 5):
                ws.merge_cells(start_row=2, start_column=col_idx, end_row=3, end_column=col_idx)

            col_widths = [24, 48, 24, 24, 24, 24, 24, 24, 24, 24, 15, 12, 12]
            for idx, width in enumerate(col_widths, 1):
                col_letter = get_column_letter(idx)
                ws.column_dimensions[col_letter].width = width

            for row_idx in range(1, 4):
                ws.row_dimensions[row_idx].height = 35

            buy_signals = self.strategy_buy_details.get(strategy_name)
            if buy_signals is None or buy_signals.empty:
                logger.warning("策略 %s 无有效买入信号" % strategy_name)
                continue

            if buy_signals is not None and not buy_signals.empty:
                buy_signals_sorted = buy_signals.sort_values(by='trade_date', ascending=False).head(100)
                detail_ws = master_wb.create_sheet(title="%s详情" % safe_name[:28])

                columns = buy_signals_sorted.columns.tolist()
                for col_idx, col_name in enumerate(columns, 1):
                    chinese_name = StrategyConfig.DETAIL_COLUMN_MAPPING.get(col_name, col_name)
                    cell = detail_ws.cell(row=1, column=col_idx, value=chinese_name)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = thin_border

                for row_idx, row in enumerate(buy_signals_sorted.itertuples(index=False), 2):
                    for col_idx, value in enumerate(row, 1):
                        cell = detail_ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.alignment = data_alignment
                        cell.border = thin_border

                for col_idx in range(1, len(columns) + 1):
                    col_letter = get_column_letter(col_idx)
                    detail_ws.column_dimensions[col_letter].width = 20

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        master_file = os.path.join(report_dir, "策略汇总报告_%s.xlsx" % timestamp)

        try:
            master_wb.save(master_file)
            logger.info("所有策略报告已汇总保存至: %s" % master_file)
            self.report_file_path = master_file
        except PermissionError:
            alt_file = os.path.join(report_dir, "策略报告_%s.xlsx" % timestamp)
            master_wb.save(alt_file)
            logger.warning("原文件被占用，报告已保存为: %s" % alt_file)
            self.report_file_path = alt_file
        except Exception as e:
            logger.error("保存策略报告出错: %s" % str(e))
            self.report_file_path = None
            return False

        logger.info("\n策略回测报告摘要:")
        for strategy_name, results in self.strategy_results.items():
            logger.info("\n策略: %s" % strategy_name)
            logger.info("买入次数: %d" % results['买入次数'])

            for period in StrategyConfig.HOLDING_PERIODS:
                win_prob = results.get('%d日上涨概率(%%)' % period, 0)
                avg_return = results.get('%d日平均收益率(%%)' % period, 0)
                std_dev = results.get('%d日标准差(%%)' % period, 0)
                logger.info("%d日 胜率: %.2f%% | 收益: %.2f%% | 波动率: %.2f%%" % (period, win_prob, avg_return, std_dev))

            # 风险调整收益
            if '总体胜率(%)' in results:
                logger.info("总体胜率: %.2f%% | 盈亏比: %.2f" % (results.get('总体胜率(%)', 0), results.get('盈亏比', 0)))

            # 计算信息比率（如果多个策略）
            if len(self.strategy_results) > 1:
                logger.info("相对表现信息请查看完整报告")

        return True

    def get_today_matching_stocks(self, strategies):
        """获取当天符合策略的股列表"""
        if self.return_data is None:
            logger.warning("没有数据用于筛选当日策略股票")
            return {}, None

        latest_date = self.return_data['date'].max()
        logger.info("获取 %s 符合策略的股票列表" % latest_date.strftime('%Y-%m-%d'))

        today_matches = {}
        all_strategies = StrategyRegistry.list_strategies()
        if 'all' in strategies:
            strategies_to_run = all_strategies
        else:
            strategies_to_run = [s for s in strategies if s in all_strategies]

        logger.info("将要运行的策略: %s" % ', '.join(strategies_to_run))

        latest_data = self.return_data[self.return_data['date'] == latest_date]

        required_cols = ['code', 'date', 'close', 'pct_chg']
        if 'name' in self.return_data.columns:
            required_cols.insert(1, 'name')

        latest_data = latest_data[required_cols]

        for strategy_name in strategies_to_run:
            strategy_func = StrategyRegistry.get_strategy(strategy_name)
            if not strategy_func:
                logger.warning("策略 %s 未注册，跳过" % strategy_name)
                continue

            try:
                strategy_data = strategy_func(self.return_data, self.dynamic_params)

                if 'buy_signal' in strategy_data.columns:
                    mask = (strategy_data['date'] == latest_date) & (strategy_data['buy_signal'] == 1)
                    if mask.any():
                        result = latest_data[mask[latest_data.index]]
                        today_matches[strategy_name] = result.sort_values('code')
                        logger.info("策略 %s 当日匹配 %d 只股票" % (strategy_name, len(today_matches[strategy_name])))
                    else:
                        today_matches[strategy_name] = pd.DataFrame(columns=required_cols)
            finally:
                if 'strategy_data' in locals():
                    del strategy_data
                gc.collect()

        return today_matches, latest_date

    def send_strategy_email(self, today_matches, date):
        """发送包含当日策略匹配股票的邮件"""
        if not today_matches:
            logger.warning("没有匹配的股票数据可发送邮件")
            return False

        date_str = date.strftime('%Y-%m-%d')

        msg = MIMEMultipart()
        msg['From'] = EmailConfig.SMTP_USER
        msg['To'] = ", ".join(EmailConfig.RECIPIENTS)
        msg['Subject'] = "%s-report" % date_str

        body_parts = ["<h3>%s 策略匹配列表</h3>" % date.strftime('%Y-%m-%d')]

        for strategy, df in today_matches.items():
            strategy_desc = StrategyRegistry.get_description(strategy)

            body_parts.append("<h4>%s（共 %d 只）</h4>" % (strategy, len(df)))
            body_parts.append("<p style='font-size:12px; color:#666;'>%s</p>" % strategy_desc)

            if not df.empty:
                chinese_columns = {col: StrategyConfig.DETAIL_COLUMN_MAPPING.get(col, col) for col in df.columns}
                df_renamed = df.rename(columns=chinese_columns)
                body_parts.append(df_renamed.to_html(index=False, justify='left', classes='table table-striped'))
            else:
                body_parts.append("<p>无匹配股票</p>")

            body_parts.append("<br>")
            del df

        body = ''.join(body_parts)
        msg.attach(MIMEText(body, 'html', 'utf-8'))

        del body_parts, body

        if self.report_file_path and os.path.exists(self.report_file_path):
            try:
                with open(self.report_file_path, "rb") as attachment:
                    part = MIMEBase("application", "octet-stream")
                    part.set_payload(attachment.read())

                encoders.encode_base64(part)
                filename = "%s-report.xlsx" % date_str
                part.add_header(
                    "Content-Disposition",
                    "attachment",
                    filename=("utf-8", "", filename)
                )
                msg.attach(part)
                logger.info("已添加附件: %s" % filename)
            except Exception as e:
                logger.error("添加邮件附件失败: %s" % str(e))
            finally:
                gc.collect()

        max_retries = 5
        for attempt in range(max_retries):
            server = None
            try:
                logger.info("尝试使用SSL连接... (尝试 %d/%d)" % (attempt + 1, max_retries))
                context = ssl.create_default_context()
                server = smtplib.SMTP_SSL(EmailConfig.SMTP_SERVER, EmailConfig.SMTP_PORT, context=context)
                logger.info("已连接到SMTP服务器，准备登录...")
                server.login(EmailConfig.SMTP_USER, EmailConfig.SMTP_PASSWORD)
                logger.info("登录成功，准备发送邮件...")
                text = msg.as_string()
                server.sendmail(EmailConfig.SMTP_USER, EmailConfig.RECIPIENTS, text)
                logger.info("策略匹配股票邮件发送成功（含附件）")
                return True
            except Exception as e:
                logger.error("发送邮件失败 (尝试 %d/%d): %s" % (attempt + 1, max_retries, str(e)))
                if attempt < max_retries - 1:
                    logger.info("等待5秒后重试...")
                    time.sleep(5)
            finally:
                if server:
                    try:
                        server.quit()
                    except Exception as e:
                        logger.debug("关闭SMTP连接时出现小问题（可忽略）: %s" % str(e))

        return False


# ----------------------------
# 交易策略实现
# ----------------------------
@StrategyRegistry.register("均线交叉策略")
def optimized_ma_crossover_strategy(df, dynamic_params=None):
    """优化的移动平均线交叉策略"""
    dynamic_params = dynamic_params or DynamicParams()

    def _calculate_macd(group):
        close = group['close']
        ema_fast = close.ewm(span=StrategyConfig.MACD_FAST, adjust=False).mean()
        ema_slow = close.ewm(span=StrategyConfig.MACD_SLOW, adjust=False).mean()
        dif = ema_fast - ema_slow
        dea = dif.ewm(span=StrategyConfig.MACD_SIGNAL, adjust=False).mean()
        macd_hist = 2 * (dif - dea)
        return dif, dea, macd_hist

    def _calculate_rsi(group):
        close = group['close']
        delta = close.diff()
        gain = delta.where(delta > 0, 0)
        loss = -delta.where(delta < 0, 0)
        avg_gain = gain.rolling(window=StrategyConfig.RSI_WINDOW, min_periods=1).mean()
        avg_loss = loss.rolling(window=StrategyConfig.RSI_WINDOW, min_periods=1).mean()
        rs = avg_gain / avg_loss.replace(0, 0.001)
        rsi = 100 - (100 / (1 + rs))
        return rsi

    def _apply_group(group):
        group = group.sort_values('trade_date')

        group['ma5'] = group['close'].rolling(window=StrategyConfig.MA_WINDOWS['short']).mean()
        group['ma20'] = group['close'].rolling(window=StrategyConfig.MA_WINDOWS['long']).mean()
        group['ma60'] = group['close'].rolling(window=60).mean()

        group['vol_ma5'] = group['volume'].rolling(window=StrategyConfig.VOL_WINDOW).mean()
        group['vol_above_ma5'] = group['volume'] > group['vol_ma5']
        group['vol_above_ma5_2d'] = group['vol_above_ma5'].rolling(2).sum() == 2

        dif, dea, macd_hist = _calculate_macd(group)
        group['macd_dif'] = dif
        group['macd_dea'] = dea
        group['macd_hist'] = macd_hist

        group['rsi'] = _calculate_rsi(group)

        group['buy_signal'] = 0
        group['sell_signal'] = 0
        group['signal_score'] = 0

        cross_up = (group['ma5'] > group['ma20']) & (group['ma5'].shift(1) <= group['ma20'].shift(1))
        group.loc[cross_up, 'signal_score'] += 3
        
        vol_condition = group['volume'] > 1.5 * group['vol_ma5']
        group.loc[vol_condition, 'signal_score'] += 2
        
        macd_positive = group['macd_hist'] > 0
        group.loc[macd_positive, 'signal_score'] += 2
        
        macd_gold = group['macd_dif'] > group['macd_dea']
        group.loc[macd_gold, 'signal_score'] += 2
        
        rsi_not_oversold = group['rsi'] > StrategyConfig.RSI_OVERSOLD
        group.loc[rsi_not_oversold, 'signal_score'] += 1.5
        
        rsi_not_overbought = group['rsi'] < StrategyConfig.RSI_OVERBOUGHT
        group.loc[rsi_not_overbought, 'signal_score'] += 1.5
        
        bullish_candle = group['close'] > group['open']
        group.loc[bullish_candle, 'signal_score'] += 1
        
        vol_continuous = group['vol_above_ma5_2d']
        group.loc[vol_continuous, 'signal_score'] += 1
        
        above_ma60 = group['close'] > group['ma60']
        group.loc[above_ma60, 'signal_score'] += 1.5

        group.loc[group['signal_score'] >= dynamic_params.MIN_SCORE_THRESHOLD, 'buy_signal'] = 1

        cross_down = (group['ma5'] < group['ma20']) & (group['ma5'].shift(1) >= group['ma20'].shift(1))
        group.loc[cross_down, 'sell_signal'] = 1

        group['trend_strength'] = (group['ma5'] - group['ma20']) / group['ma20'] * 100

        return group

    return df.groupby(df['code'], group_keys=False, as_index=False).apply(_apply_group)

@StrategyRegistry.register("成交量突破策略")
def volume_breakout_strategy(df, dynamic_params=None):
    """成交量突破策略"""
    dynamic_params = dynamic_params or DynamicParams()

    def _apply_group(group):
        group = group.sort_values('trade_date')
        group['buy_signal'] = 0
        group['signal_score'] = 0
        
        group['vol_mean'] = group['volume'].rolling(window=StrategyConfig.MA_WINDOWS['long']).mean()
        group['vol_std'] = group['volume'].rolling(window=StrategyConfig.MA_WINDOWS['long']).std()
        
        base_breakout = group['volume'] > group['vol_mean'] + StrategyConfig.VOLUME_THRESHOLDS['breakout'] * group['vol_std']
        group.loc[base_breakout, 'signal_score'] += 4
        
        group.loc[base_breakout & (group['close'] > group['open']), 'signal_score'] += 2
        
        group['high_20d'] = group['high'].rolling(window=20).max()
        group.loc[base_breakout & (group['close'] > group['high_20d'].shift(1)), 'signal_score'] += 3
        
        group.loc[base_breakout & (group['volume'] > 1.5 * group['volume'].shift(1)), 'signal_score'] += 2
        
        if 'rsi' in group.columns:
            group.loc[base_breakout & (group['rsi'].between(30, 70)), 'signal_score'] += 2
        
        group.loc[group['signal_score'] >= dynamic_params.MIN_SCORE_THRESHOLD, 'buy_signal'] = 1
        
        return group

    return df.groupby(df['code'], group_keys=False, as_index=False).apply(_apply_group)


@StrategyRegistry.register("多均线共振策略")
def multi_ma_resonance_strategy(df, dynamic_params=None):
    """多均线共振策略"""
    dynamic_params = dynamic_params or DynamicParams()

    def _calculate_macd(group):
        close = group['close']
        ema_fast = close.ewm(span=14, adjust=False).mean()
        ema_slow = close.ewm(span=53, adjust=False).mean()
        dif = ema_fast - ema_slow
        dea = dif.ewm(span=5, adjust=False).mean()
        macd_hist = 2 * (dif - dea)
        return dif, dea, macd_hist

    def _calculate_bollinger(group, window=20, num_std=2):
        close = group['close']
        ma = close.rolling(window=window).mean()
        std = close.rolling(window=window).std()
        upper = ma + num_std * std
        lower = ma - num_std * std
        return upper, lower, ma

    def _calculate_rsi_lines(group):
        close = group['close']
        delta = close.diff()
        gain = delta.where(delta > 0, 0)
        loss = -delta.where(delta < 0, 0)

        rsi6 = 100 - (100 / (1 + gain.rolling(6).mean() / loss.rolling(6).mean().replace(0, 0.001)))
        rsi12 = 100 - (100 / (1 + gain.rolling(12).mean() / loss.rolling(12).mean().replace(0, 0.001)))
        rsi24 = 100 - (100 / (1 + gain.rolling(24).mean() / loss.rolling(24).mean().replace(0, 0.001)))
        return rsi6, rsi12, rsi24

    def _apply_group(group):
        group = group.sort_values('trade_date')
        group['signal_score'] = 0

        group['ma5'] = group['close'].rolling(window=5).mean()
        group['ma10'] = group['close'].rolling(window=10).mean()
        group['ma20'] = group['close'].rolling(window=20).mean()
        group['ma60'] = group['close'].rolling(window=60).mean()
        group['ma90'] = group['close'].rolling(window=90).mean()
        group['ma120'] = group['close'].rolling(window=120).mean()

        for ma in ['ma5', 'ma10', 'ma20', 'ma60', 'ma90', 'ma120']:
            cond = group[ma] > group[ma].shift(1)
            group.loc[cond, 'signal_score'] += 1

        order_conditions = [
            (group['ma5'] > group['ma10']),
            (group['ma10'] > group['ma20']),
            (group['ma20'] > group['ma60']),
            (group['ma60'] > group['ma90']),
            (group['ma90'] > group['ma120'])
        ]
        for cond in order_conditions:
            group.loc[cond, 'signal_score'] += 1

        ma_cross = (group['ma5'] > group['ma10']) & (group['ma5'].shift(1) <= group['ma10'].shift(1))
        group.loc[ma_cross, 'signal_score'] += 3

        group['vol_ma20'] = group['volume'].rolling(window=20).mean()
        vol_increase = group['volume'] > group['vol_ma20']
        group.loc[vol_increase, 'signal_score'] += 2

        dif, dea, macd_hist = _calculate_macd(group)
        group['macd_dif'] = dif
        group['macd_dea'] = dea
        group['macd_hist'] = macd_hist

        macd_cross = (dif > dea) & (dif.shift(1) <= dea.shift(1))
        macd_turn_red = (macd_hist > 0) & (macd_hist.shift(1) <= 0)
        group.loc[macd_cross, 'signal_score'] += 3
        group.loc[macd_turn_red, 'signal_score'] += 2

        upper, lower, boll_ma = _calculate_bollinger(group)
        group['boll_upper'] = upper
        group['boll_lower'] = lower
        group['boll_width'] = upper - lower
        boll_just_opened = (
                (group['boll_width'] > group['boll_width'].shift(1)) &
                (group['boll_width'].shift(1) < group['boll_width'].shift(2))
        )
        group.loc[boll_just_opened, 'signal_score'] += 2

        rsi6, rsi12, rsi24 = _calculate_rsi_lines(group)
        group['rsi6'] = rsi6
        group['rsi12'] = rsi12
        group['rsi24'] = rsi24

        rsi_up = (
                (rsi6 > rsi6.shift(1)) &
                (rsi12 > rsi12.shift(1)) &
                (rsi24 > rsi24.shift(1)) &
                (rsi6 < StrategyConfig.RSI_OVERBOUGHT) &
                (rsi12 < StrategyConfig.RSI_OVERBOUGHT) &
                (rsi24 < StrategyConfig.RSI_OVERBOUGHT)
        )
        group.loc[rsi_up, 'signal_score'] += 3

        group['buy_signal'] = 0
        group.loc[group['signal_score'] >= dynamic_params.MIN_SCORE_THRESHOLD, 'buy_signal'] = 1

        return group

    return df.groupby('code', group_keys=False).apply(_apply_group)


@StrategyRegistry.register("神奇九转策略")
def optimized_wonderful_9_turn_strategy(df, dynamic_params=None):
    """优化的神奇九转策略"""
    dynamic_params = dynamic_params or DynamicParams()

    def _wonderful_9_turn_apply_group(group):
        group = group.sort_values('trade_date').copy()
        group['signal_score'] = 0

        close_lt_4days = group['close'] < group['close'].shift(4)
        streak_count = close_lt_4days.rolling(window=9).sum()

        low_8_lt_6 = group['low'].shift(1) < group['low'].shift(3)
        low_8_lt_7 = group['low'].shift(1) < group['low'].shift(2)
        low_9_lt_6 = group['low'] < group['low'].shift(3)
        low_9_lt_7 = group['low'] < group['low'].shift(2)
        low_condition = low_8_lt_6 | low_8_lt_7 | low_9_lt_6 | low_9_lt_7

        base_signal = (streak_count == 9) & low_condition
        group.loc[base_signal, 'signal_score'] += 4

        delta = group['close'].diff()
        gain = delta.where(delta > 0, 0)
        loss = -delta.where(delta < 0, 0)
        avg_gain = gain.rolling(window=StrategyConfig.RSI_WINDOW, min_periods=1).mean()
        avg_loss = loss.rolling(window=StrategyConfig.RSI_WINDOW, min_periods=1).mean()
        rs = avg_gain / avg_loss.replace(0, 0.001)
        group['rsi'] = 100 - (100 / (1 + rs))
        
        rsi_condition = group['rsi'] < getattr(StrategyConfig, 'RSI_OVERSOLD', 40)
        group.loc[base_signal & rsi_condition, 'signal_score'] += 2

        ema_fast = group['close'].ewm(span=StrategyConfig.MACD_FAST, adjust=False).mean()
        ema_slow = group['close'].ewm(span=StrategyConfig.MACD_SLOW, adjust=False).mean()
        dif = ema_fast - ema_slow
        dea = dif.ewm(span=StrategyConfig.MACD_SIGNAL, adjust=False).mean()
        group['macd_hist'] = 2 * (dif - dea)
        
        macd_improving = (group['macd_hist'] > group['macd_hist'].shift(1))
        macd_negative = group['macd_hist'] < 0
        macd_condition = macd_improving & macd_negative
        group.loc[base_signal & macd_condition, 'signal_score'] += 2

        group['vol_ma20'] = group['volume'].rolling(window=20).mean()
        volume_condition = group['volume'] > StrategyConfig.VOLUME_THRESHOLDS['resonance'] * group['vol_ma20']
        group.loc[base_signal & volume_condition, 'signal_score'] += 2

        group['ma20'] = group['close'].rolling(window=20).mean()
        group['ma60'] = group['close'].rolling(window=60).mean()
        trend_up = (group['ma20'] > group['ma60']) & \
                   (group['ma20'] > group['ma20'].shift(1)) & \
                   (group['ma60'] > group['ma60'].shift(1))
        group.loc[base_signal & trend_up, 'signal_score'] += 2

        price_near_ma = group['close'] >= getattr(StrategyConfig, 'MA_SUPPORT_RATIO', 0.98) * group['ma20']
        group.loc[base_signal & price_near_ma, 'signal_score'] += 2

        group['buy_signal'] = 0
        final_signal = group['signal_score'] >= StrategyConfig.MIN_SCORE_THRESHOLD
        final_signal = final_signal & (group['close'] > group['open'])

        group.loc[final_signal, 'buy_signal'] = 1
        group['signal_strength'] = group['signal_score'] / 14

        return group

    return df.groupby('code', group_keys=False, observed=True).apply(_wonderful_9_turn_apply_group)

@StrategyRegistry.register("涨停回调策略")
def limit_up_pullback_strategy(df, dynamic_params=None):
    """涨停回调策略"""
    dynamic_params = dynamic_params or DynamicParams()

    def _apply_group(group):
        group = group.sort_values('trade_date').copy()
        group['buy_signal'] = 0
        group['signal_score'] = 0

        limit_up_cond = (group['pct_chg'] >= StrategyConfig.LIMIT_UP_PCT_THRESHOLD)
        group['is_limit_up'] = limit_up_cond

        group['has_limit_up_90d'] = group['is_limit_up'].rolling(window=90, min_periods=1).sum().shift(1, fill_value=0) > 0
        group['first_limit_up_90d'] = limit_up_cond & (~group['has_limit_up_90d'])
        group.loc[group['first_limit_up_90d'], 'signal_score'] += 3

        group['high_90d'] = group['high'].rolling(window=90, min_periods=1).max()
        group['valid_limit_up'] = group['first_limit_up_90d'] & (group['high'] >= group['high_90d'])
        group.loc[group['valid_limit_up'], 'signal_score'] += 3

        group['limit_up_price'] = np.nan
        group['pre_limit_high'] = np.nan
        group['limit_up_volume'] = np.nan
        group['limit_up_date'] = pd.NaT

        valid_mask = group['valid_limit_up']
        if not valid_mask.empty:
            group.loc[valid_mask, 'limit_up_date'] = pd.to_datetime(group.loc[valid_mask, 'trade_date'])

        group.loc[valid_mask, 'limit_up_price'] = group.loc[valid_mask, 'high']
        group.loc[valid_mask, 'limit_up_volume'] = group.loc[valid_mask, 'volume']
        group.loc[valid_mask, 'limit_up_date'] = group.loc[valid_mask, 'trade_date']
        group.loc[valid_mask, 'pre_limit_high'] = group.loc[valid_mask, 'high'].shift(1).rolling(window=89, min_periods=1).max()

        group['limit_up_price'] = group['limit_up_price'].ffill()
        group['pre_limit_high'] = group['pre_limit_high'].ffill()
        group['limit_up_volume'] = group['limit_up_volume'].ffill()
        group['limit_up_date'] = group['limit_up_date'].ffill()

        group['pullback_low'] = group['pre_limit_high'] * 0.98
        group['pullback_high'] = group['pre_limit_high'] * 1.02

        group['is_shrinking_volume'] = group['volume'] < group['limit_up_volume'] * 0.5
        group.loc[group['is_shrinking_volume'], 'signal_score'] += 1

        has_valid_limit_up = ~group['limit_up_price'].isna()
        in_pullback_range = (group['low'] >= group['pullback_low']) & (group['high'] <= group['pullback_high'])
        group['in_pullback_zone'] = has_valid_limit_up & in_pullback_range & group['is_shrinking_volume']
        group.loc[group['in_pullback_zone'], 'signal_score'] += 2

        group['pullback_count'] = 0
        pullback_counter = 0
        last_limit_up_idx = None

        for i in range(len(group)):
            if group['valid_limit_up'].iloc[i]:
                pullback_counter = 0
                last_limit_up_idx = i
                continue

            if last_limit_up_idx is None:
                continue

            if i <= last_limit_up_idx:
                continue

            current_in = group['in_pullback_zone'].iloc[i]
            prev_in = group['in_pullback_zone'].iloc[i - 1] if i > 0 else False

            if current_in and not prev_in:
                pullback_counter += 1

            group.iloc[i, group.columns.get_loc('pullback_count')] = pullback_counter
            
        group.loc[group['pullback_count'] >= 2, 'signal_score'] += 2
        group.loc[group['pullback_count'] >= 3, 'signal_score'] += 2

        group['vol_ma5'] = group['volume'].rolling(window=5, min_periods=1).mean()
        group['is_rising_volume'] = (group['volume'] > group['vol_ma5'] * 1.5) & (group['pct_chg'] > 0)
        group.loc[group['is_rising_volume'], 'signal_score'] += 3

        buy_cond = (
                has_valid_limit_up &
                (group['pullback_count'] >= 2) &
                group['is_rising_volume']
        )

        group.loc[buy_cond & (group['signal_score'] >= dynamic_params.MIN_SCORE_THRESHOLD), 'buy_signal'] = 1

        temp_cols = ['is_limit_up', 'has_limit_up_90d', 'high_90d', 'valid_limit_up',
                     'limit_up_price', 'limit_up_volume', 'limit_up_date', 'pre_limit_high',
                     'pullback_low', 'pullback_high', 'is_shrinking_volume',
                     'in_pullback_zone', 'vol_ma5', 'is_rising_volume']
        existing_cols = [col for col in temp_cols if col in group.columns]
        if existing_cols:
            group = group.drop(columns=existing_cols)

        return group

    return df.groupby('code', group_keys=False, as_index=False).apply(_apply_group)


@StrategyRegistry.register("孕阳线策略")
def big_red_after_black_strategy(df, dynamic_params=None):
    """孕阳线策略"""
    dynamic_params = dynamic_params or DynamicParams()

    def _apply_group(group):
        group = group.sort_values('trade_date')
        group['buy_signal'] = 0
        group['signal_score'] = 0

        prev_pct = group['pct_chg'].shift(1)
        prev_open = group['open'].shift(1)
        prev_close = group['close'].shift(1)
        body_pct = abs(prev_close - prev_open) / prev_open * 100
        big_black = (prev_pct <= -7) & (body_pct >= 70)
        group.loc[big_black, 'signal_score'] += 5

        current_bullish = group['close'] > group['open']
        price_in_range = (group['open'] > prev_close) & (group['close'] < prev_open)
        group.loc[big_black & current_bullish & price_in_range, 'signal_score'] += 5
        
        vol_shrink = group['volume'] < group['volume'].shift(1)
        group.loc[big_black & current_bullish & price_in_range & vol_shrink, 'signal_score'] += 3

        group.loc[group['signal_score'] >= dynamic_params.MIN_SCORE_THRESHOLD, 'buy_signal'] = 1
        return group

    return df.groupby(df['code'], group_keys=False, as_index=False).apply(_apply_group)


@StrategyRegistry.register("组合策略")
def ensemble_strategy(df, dynamic_params=None):
    """多策略共振策略 - 多个策略同时触发时才买入"""
    dynamic_params = dynamic_params or DynamicParams()

    base_strategies = [
        ('均线交叉策略', optimized_ma_crossover_strategy),
        ('成交量突破策略', volume_breakout_strategy),
        ('多均线共振策略', multi_ma_resonance_strategy)
    ]

    def _apply_group(group):
        group = group.sort_values('trade_date')
        group['signal_count'] = 0
        group['buy_signal'] = 0
        group['signal_score'] = 0

        for strat_name, strat_func in base_strategies:
            try:
                temp_group = strat_func(pd.DataFrame(group), dynamic_params)
                if 'buy_signal' in temp_group.columns:
                    group['signal_count'] += temp_group['buy_signal']
                    if 'signal_score' in temp_group.columns:
                        group['signal_score'] += temp_group['signal_score']
            except Exception as e:
                logger.warning(f"组合策略中执行 {strat_name} 失败: {e}")

        # 至少2个策略同时触发且达到最低分数才买入
        group.loc[
            (group['signal_count'] >= 2) &
            (group['signal_score'] >= dynamic_params.MIN_SCORE_THRESHOLD),
            'buy_signal'
        ] = 1

        group.loc[group['buy_signal'] == 1, 'signal_score'] = min(group['signal_score'] + 5, 20)

        temp_cols = ['signal_count']
        existing_cols = [col for col in temp_cols if col in group.columns]
        if existing_cols:
            group = group.drop(columns=existing_cols)

        return group

    return df.groupby('code', group_keys=False, as_index=False).apply(_apply_group)


# ----------------------------
# 主函数
# ----------------------------
def main():
    parser = argparse.ArgumentParser(description='股票策略回测系统（优化版）')
    parser.add_argument('--parallel', action='store_true', help='是否启用并行处理')
    parser.add_argument('--pool_size', type=int, default=2, help='数据库连接池大小')
    parser.add_argument('--start_date', type=str, required=True, help='回测开始日期 (YYYY-MM-DD)')
    parser.add_argument('--end_date', type=str, help='回测结束日期 (YYYY-MM-DD), 默认为今天')
    parser.add_argument('--strategies', nargs='+', default=['all'], help='要回测的策略列表，默认全部')
    parser.add_argument('--send_email', action='store_true', help='是否发送邮件报告')
    parser.add_argument('--save_db', action='store_true', default=True, help='是否保存结果到数据库')
    parser.add_argument('--disable_ml', action='store_true', help='禁用机器学习信号过滤')
    parser.add_argument('--disable_stop_loss', action='store_true', help='禁用止损止盈')
    parser.add_argument('--disable_transaction_cost', action='store_true', help='禁用交易成本计算')

    args = parser.parse_args()

    db_params = {
        'dbname': DbConfig.DB_NAME,
        'user': DbConfig.DB_USER,
        'password': DbConfig.DB_PASSWORD,
        'host': DbConfig.DB_HOST,
        'port': DbConfig.DB_PORT
    }

    backtest = StockStrategyBacktest(db_params, pool_size=args.pool_size)

    try:
        end_date = args.end_date or datetime.now().strftime('%Y-%m-%d')
        stock_data = backtest.fetch_stock_data(args.start_date, end_date)
        if stock_data is None:
            logger.error("无法获取股票数据，程序退出")
            return

        return_data = backtest.calculate_returns()
        if return_data is None:
            logger.error("无法计算收益率，程序退出")
            return

        if 'all' in args.strategies:
            strategies_to_run = StrategyRegistry.list_strategies()
        else:
            strategies_to_run = [s for s in args.strategies if s in StrategyRegistry.list_strategies()]

        logger.info("即将回测的策略: %s" % ', '.join(strategies_to_run))

        if args.parallel and len(strategies_to_run) > 1:
            with concurrent.futures.ThreadPoolExecutor(max_workers=min(5, len(strategies_to_run))) as executor:
                futures = []
                for strategy_name in strategies_to_run:
                    strategy_func = StrategyRegistry.get_strategy(strategy_name)
                    if strategy_func:
                        futures.append(executor.submit(
                            backtest.apply_strategy,
                            return_data,
                            strategy_func,
                            strategy_name,
                            ml_filter=not args.disable_ml,
                            enable_stop_loss=not args.disable_stop_loss,
                            enable_transaction_cost=not args.disable_transaction_cost
                        ))

                for future in concurrent.futures.as_completed(futures):
                    try:
                        future.result()
                    except Exception as e:
                        logger.error("并行处理策略时出错: %s" % str(e))
        else:
            for strategy_name in strategies_to_run:
                strategy_func = StrategyRegistry.get_strategy(strategy_name)
                if strategy_func:
                    logger.info("开始回测策略: %s" % strategy_name)
                    backtest.apply_strategy(
                        return_data,
                        strategy_func,
                        strategy_name,
                        ml_filter=not args.disable_ml,
                        enable_stop_loss=not args.disable_stop_loss,
                        enable_transaction_cost=not args.disable_transaction_cost
                    )

        if args.save_db:
            backtest.save_results_to_db(args.start_date, end_date)

        backtest.generate_report(start_date=args.start_date, end_date=end_date)

        if args.send_email:
            today_matches, latest_date = backtest.get_today_matching_stocks(args.strategies)
            backtest.send_strategy_email(today_matches, latest_date)

    except Exception as e:
        logger.error("回测过程中发生错误: %s" % str(e))
        logger.error(traceback.format_exc())
    finally:
        backtest.close_pool()

if __name__ == "__main__":
    main()