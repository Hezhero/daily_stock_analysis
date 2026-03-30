# -*- coding: utf-8 -*-
import time
import argparse
import concurrent.futures
import subprocess
import gc
import logging
import os
import traceback
import ssl
import json
from email.mime.base import MIMEBase
from email import encoders
import baostock as bs
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from psycopg2 import sql
from psycopg2.pool import SimpleConnectionPool

# cd baostock && conda activate tushare && python strategy.py --parallel --send_email --pool_size 5 --start_date 2005-01-01 --strategies 多均线共振策略 均线交叉策略 孕阳线策略 神奇九转策略 成交量突破策略 涨停回调策略 连续平稳后涨停策略
# ----------------------------
# 常量定义和配置
# ----------------------------

"""邮件发送配置"""
class EmailConfig:
    SMTP_SERVER = "smtp.qq.com"  # 例如: smtp.qq.com 或 smtp.gmail.com
    SMTP_PORT = 465  # ssl 通常为465
    SMTP_USER = "851448443@qq.com"
    SMTP_PASSWORD = "aofwlgcsobymbfdj"  # 部分邮箱需要应用专用密码
    RECIPIENTS = ["la9408531@163.com", "1049220782@qq.com","122755347@qq.com"]

"""数据库配置"""
class DbConfig:
    DB_HOST = "localhost"
    DB_PORT = "5431"
    DB_USER = "root"
    DB_PASSWORD = "123629He"
    DB_NAME = "baostock"

class StrategyConfig:
    """策略参数集中管理"""
    LIMIT_UP_PCT_THRESHOLD = 10  # 涨停阈值(%)
    MA_WINDOWS = {
        'short': 5,  # 短期均线窗口
        'long': 20  # 长期均线窗口
    }
    HOLDING_PERIODS = [1, 3, 5, 10, 20]  # 持有期天数
    VOLUME_THRESHOLDS = {
        'breakout': 2.0,  # 成交量突破标准差倍数
        'resonance': 1.5  # 量价共振放量倍数
    }
    STABILITY_THRESHOLD = 5  # 平稳策略的涨跌幅阈值(%)
    REPORT_COLUMNS = ['策略名称', '策略条件描述', '回测日期范围', '买入次数', '表现',
                      '1 日', '3 日', '5 日', '10 日', '20 日']
    # 可配置参数
    VOL_WINDOW = 5  # 成交量平均窗口 观察 1周的资金流入流出节奏
    RSI_WINDOW = 6  # RSI计算窗口 （更灵敏，能快速反映 1-3 天的强弱变化）
    MACD_FAST = 14  # MACD快速EMA周期 持仓周期（通常 1-5 个交易日）
    MACD_SLOW = 53  # MACD慢速EMA周期 持仓周期（通常 1-5 个交易日）
    MACD_SIGNAL = 5  # MACD信号线周期 持仓周期（通常 1-5 个交易日）
    MA_SUPPORT_RATIO = 0.98
    MIN_CONDITIONS = 3  # 最少需要满足的条件数
    RSI_OVERBOUGHT = 70  # RSI超买阈值
    RSI_OVERSOLD = 30  # RSI超卖阈值

    # 详情表列名中文映射
    DETAIL_COLUMN_MAPPING = {
        'code': '股票代码',
        'name': '股票名称',
        'date': '交易日期',
        'close': '收盘价',
        'pct_chg': '涨跌幅(%)',
        'open': '开盘价',
        'high': '最高价',
        'low': '最低价',
        'volume': '成交量',
        'amount': '成交额',
        'turn': '换手率',
        'buy_signal': '买入信号',
        'return_1d': '1日收益率',
        'return_3d': '3日收益率',
        'return_5d': '5日收益率',
        'return_10d': '10日收益率',
        'return_20d': '20日收益率'
        # 可以根据实际存在的列继续补充
    }

    # 策略描述映射
    DESCRIPTIONS = {
        '均线交叉策略': '5日均线上穿20日均线时收阳线，当日成交量大于5日均量50%且连续2日大于5日均量确认资金入场，MACD（6,13,5）柱状线为正且DIF>DEA确认多头趋势,RSI(6)脱离超卖区(>30)但未超买(<70),收盘价在60日线上方，产生买入信号',
        '成交量突破策略': '成交量超过20日均值2倍标准差时产生买入信号',
        '神奇九转策略': '1、满足神奇九转形态(9天内满足低点比较条件,且连续9天收盘价低于4天前的收盘价)2、RSI低于超卖阈值 3、MACD绿柱缩短且仍在0轴下方 4、成交量放大 5、20日均线和60日均线齐头向上，且收盘价接近20日均线获得支撑，收盘确认',
        'N字反包策略': '回调后的收盘价低于前一个上涨波段的高点，且当前收盘价高于该高点时产生买入信号',
        '涨停回调策略': '股票涨停后，后续回调不跌破涨停板的最低价时产生买入信号',
        '连续平稳后涨停策略': '连续10个交易日涨跌幅在{}%以内，第11个交易日涨停时买入'.format(STABILITY_THRESHOLD),
        'MACD月线金叉+20日线策略': '月线MACD金叉且20日线附近放量突破时买入',
        '低位涨停换手率策略': '下跌趋势中止跌后价量齐升时买入',
        '涨停回调量价共振策略': '涨停后缩量且20日均线上行时的特定K线形态买入',
        '孕阳线策略': '上个交易日为跌幅≥7%且实体占比≥70%的大阴线，本交易日开盘价和收盘价均位于前一日开盘价与收盘价之间，当日开盘价至少高开2%且收阳线时产生买入信号',
        '多均线共振策略': '1、120、90、60、20日这4条均线都齐头向上，5日均线刚好金叉10日均线2、成交量放大3、MACD（14,53,5）金叉，且量柱由绿翻红4、boll带开口5、RSI三线齐头向上'
    }


# ----------------------------
# 策略注册机制
# ----------------------------
class StrategyRegistry:
    """策略注册器，集中管理所有交易策略"""
    _strategies = {}

    @classmethod
    def register(cls, name, description=""):
        """装饰器函数，用于注册策略"""

        def decorator(func):
            cls._strategies[name] = {
                'function': func,
                'description': description or StrategyConfig.DESCRIPTIONS.get(name, "")
            }
            return func

        return decorator

    @classmethod
    def get_strategy(cls, name):
        """获取指定名称的策略函数"""
        return cls._strategies.get(name, {}).get('function')

    @classmethod
    def get_description(cls, name):
        """获取指定策略的描述"""
        return cls._strategies.get(name, {}).get('description', "")

    @classmethod
    def list_strategies(cls):
        """列出所有注册的策略"""
        return list(cls._strategies.keys())


# ----------------------------
# 日志配置
# ----------------------------
def setup_logger():
    """配置日志系统"""
    logger = logging.getLogger('strategy_backtest')
    logger.setLevel(logging.DEBUG)

    # 创建日志目录
    log_dir = "./result"
    os.makedirs(log_dir, exist_ok=True)

    # 文件处理器（记录所有级别日志）
    file_handler = logging.FileHandler(
        os.path.join(log_dir, "backtest.log"),
        encoding='utf-8'
    )
    file_handler.setLevel(logging.DEBUG)

    # 控制台处理器（只记录INFO及以上级别）
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)

    # 格式化器
    formatter = logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)

    # 添加处理器
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    return logger


logger = setup_logger()

def apply_strategy_independent(stock_data, strategy_func, strategy_name="策略"):
    """独立于数据库连接的应用策略函数 - 优化版"""
    if stock_data is None or stock_data.empty:
        logger.warning("{}: 没有数据用于策略评估".format(strategy_name))
        return None, None

    try:
        # 分块处理数据，避免内存溢出
        logger.info("开始应用策略: {}".format(strategy_name))

        # 按股票代码分组处理，避免一次性加载所有数据
        codes = stock_data['code'].unique()
        all_buy_signals = []
        all_results = []

        # 分批处理股票数据
        batch_size = 100  # 每批处理100只股票
        for i in range(0, len(codes), batch_size):
            batch_codes = codes[i:i + batch_size]
            batch_data = stock_data[stock_data['code'].isin(batch_codes)].copy()

            # 应用策略
            strategy_data = strategy_func(batch_data)

            if 'buy_signal' in strategy_data.columns:
                buy_signals = strategy_data[strategy_data['buy_signal'] == 1].copy()
                if not buy_signals.empty:
                    all_buy_signals.append(buy_signals)

            # 及时释放内存
            del batch_data, strategy_data
            gc.collect()

        if all_buy_signals:
            buy_signals_combined = pd.concat(all_buy_signals, ignore_index=True)
        else:
            buy_signals_combined = pd.DataFrame()

        logger.info("{}: 生成 {} 个买入信号".format(strategy_name, len(buy_signals_combined)))

        if buy_signals_combined.empty:
            logger.warning("{}: 没有有效的买入信号".format(strategy_name))
            return None, None

        # 计算不同持有期的表现
        results = {
            '策略名称': strategy_name,
            '买入次数': len(buy_signals_combined),
            '策略描述': StrategyConfig.DESCRIPTIONS.get(strategy_name, "")
        }

        for period in StrategyConfig.HOLDING_PERIODS:
            return_col = 'return_{}d'.format(period)
            if return_col not in buy_signals_combined.columns:
                logger.warning("{}: 缺少{}列，跳过该持有期".format(strategy_name, return_col))
                continue

            # 处理无效值
            returns_series = buy_signals_combined[return_col]
            valid_returns = returns_series.replace([np.inf, -np.inf], np.nan).dropna()

            if valid_returns.empty:
                results['{}日上涨概率(%)'.format(period)] = 0
                results['{}日平均收益率(%)'.format(period)] = 0
                continue

            # 计算正收益率比例
            pos_returns = valid_returns[valid_returns > 0]
            win_prob = len(pos_returns) / len(valid_returns) * 100

            # 计算平均收益率
            try:
                avg_return = valid_returns.mean() * 100
            except Exception as e:
                logger.warning("计算平均收益率出错: {}".format(e))
                avg_return = 0

            results['{}日上涨概率(%)'.format(period)] = win_prob
            results['{}日平均收益率(%)'.format(period)] = avg_return

        return buy_signals_combined, results

    except Exception as e:
        logger.error("应用策略 {} 时出错: {}".format(strategy_name, str(e)))
        logger.error(traceback.format_exc())
        return None, None


class StockStrategyBacktest:
    def __init__(self, db_params, pool_size=5):
        """初始化数据库连接池和结果存储"""
        self.db_params = db_params
        self.pool_size = pool_size
        self.connection_pool = None
        self.strategy_results = {}
        self.strategy_buy_details = {}
        self.stock_data = None
        self.return_data = None
        self.report_file_path = None  # 新增：存储报告文件路径

    def create_connection_pool(self):
        """创建数据库连接池"""
        if self.connection_pool is None:
            try:
                self.connection_pool = SimpleConnectionPool(
                    minconn=1,
                    maxconn=self.pool_size,
                    **self.db_params
                )
                logger.info("成功创建数据库连接池 (大小: {})".format(self.pool_size))
                return True
            except Exception as e:
                logger.error("创建数据库连接池失败: {}".format(e))
                return False
        return True

    def get_connection(self):
        """从连接池获取数据库连接"""
        if self.connection_pool is None:
            self.create_connection_pool()
        return self.connection_pool.getconn()

    def release_connection(self, conn):
        """释放数据库连接回连接池"""
        if conn:
            self.connection_pool.putconn(conn)

    def close_pool(self):
        """关闭数据库连接池"""
        if self.connection_pool:
            self.connection_pool.closeall()
            logger.info("数据库连接池已关闭")

    # 优化 fetch_stock_data 方法
    def fetch_stock_data(self, start_date, end_date, chunk_size=10000):
        """从数据库获取股票数据（分块读取优化内存）"""
        if self.stock_data is not None:
            return self.stock_data

        if not self.create_connection_pool():
            return None

        conn = self.get_connection()
        try:
            with conn.cursor() as cursor:
                query = sql.SQL("""
                                SELECT code,
                                       COALESCE(name, '') as name,
                                       date, 
                                       COALESCE(open, 0) as open, 
                                       COALESCE(high, 0) as high, 
                                       COALESCE(low, 0) as low, 
                                       COALESCE(close, 0) as close, 
                                       COALESCE(pct_chg, 0) as pct_chg, 
                                       COALESCE(volume, 0) as volume, 
                                       COALESCE(amount, 0) as amount, 
                                       COALESCE(turn, 0) as turn
                                FROM baostock_daily_history
                                WHERE date BETWEEN %s
                                  AND %s
                                ORDER BY code, date
                                """)
                cursor.execute(query, (start_date, end_date))

                # 分块读取数据
                chunks = []
                while True:
                    data = cursor.fetchmany(chunk_size)
                    if not data:
                        break
                    columns = [desc[0] for desc in cursor.description]

                    # 指定每列的数据类型以减少内存占用
                    dtype_mapping = {
                        'code': 'category',
                        'name': 'category',
                        'date': 'datetime64[ns]',
                        'open': 'float32',
                        'high': 'float32',
                        'low': 'float32',
                        'close': 'float32',
                        'pct_chg': 'float32',
                        'volume': 'int64',  # 从 int64 改为 int32
                        'amount': 'float32',  # 从 float64 改为 float32
                        'turn': 'float32'
                    }

                    df_chunk = pd.DataFrame(data, columns=columns)
                    for col, dtype in dtype_mapping.items():
                        if col in df_chunk.columns:
                            df_chunk[col] = df_chunk[col].astype(dtype)

                    chunks.append(df_chunk)
                    del df_chunk  # 处理完每个chunk后立即删除

                if chunks:
                    df = pd.concat(chunks, ignore_index=True)
                    df['date'] = pd.to_datetime(df['date'])
                    logger.info("成功获取 {} 条股票数据 ({} 只股票)".format(len(df), len(df['code'].unique())))
                    self.stock_data = df
                    return df
                else:
                    logger.warning("未获取到任何股票数据")
                    return None
        except Exception as e:
            logger.error("获取股票数据失败: {}".format(e))
            return None
        finally:
            self.release_connection(conn)

    def calculate_returns(self, stock_data=None, holding_periods=StrategyConfig.HOLDING_PERIODS):
        """计算不同持有期的收益率（向量化优化）"""
        if self.return_data is not None:
            return self.return_data

        df = stock_data if stock_data is not None else self.stock_data
        if df is None or df.empty:
            logger.warning("没有数据用于计算收益率")
            return None

        # 确保数据按股票代码和日期排序
        df = df.sort_values(['code', 'date'])

        # 转换close列为数值类型
        df['close'] = pd.to_numeric(df['close'], errors='coerce')

        # 计算不同持有期的收益率
        for period in holding_periods:
            df['return_{}d'.format(period)] = df.groupby(df['code'])['close'].pct_change(periods=period).shift(-period)

        # 删除不再需要的列以释放内存
        cols_to_drop = ['ma5', 'ma20', 'vol_mean', 'vol_std', 'lower_than_4days_ago', 'count']
        existing_cols = [col for col in cols_to_drop if col in df.columns]
        if existing_cols:
            df = df.drop(columns=existing_cols)

        logger.info("收益率计算完成，涵盖 {} 只股票".format(len(df['code'].unique())))
        self.return_data = df
        return df

    # 修改 StockStrategyBacktest 类中的 apply_strategy 方法
    def apply_strategy(self, stock_data, strategy_func, strategy_name="策略"):
        """应用交易策略并评估结果"""
        buy_signals, results = apply_strategy_independent(stock_data, strategy_func, strategy_name)

        if results is not None:
            self.strategy_results[strategy_name] = results

        if buy_signals is not None:
            self.strategy_buy_details[strategy_name] = buy_signals

        return stock_data, buy_signals

    # 优化 generate_report 方法 详情页只保留前100条数据
    def generate_report(self, output_format='excel', output_file='strategy_report', start_date='', end_date=''):
        """生成回测报告（优化格式和内容）"""
        if not self.strategy_results:
            logger.warning("没有策略回测结果可用于生成报告")
            return False

        # 确保报告目录存在
        report_dir = "./result"
        os.makedirs(report_dir, exist_ok=True)

        # 边框样式
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # 表头样式
        header_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')

        # 数据样式
        data_alignment = Alignment(horizontal='center', vertical='center')

        # 创建主工作簿
        master_wb = Workbook()
        master_wb.remove(master_wb.active)  # 移除默认sheet

        # 为每个策略创建工作表
        for strategy_name, results in self.strategy_results.items():
            # 创建策略表现工作表
            safe_name = strategy_name[:31]  # Excel工作表名称最大31字符
            ws = master_wb.create_sheet(title="{}表现".format(safe_name))

            # 添加表头
            for col_idx, col_name in enumerate(StrategyConfig.REPORT_COLUMNS, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            # 构建报告数据
            date_range = '{} 至 {}'.format(start_date, end_date)
            row_data = [
                strategy_name,
                results.get('策略描述', ""),
                date_range,
                results['买入次数'],
                '上涨概率',
                results.get('1日上涨概率(%)', 0),
                results.get('3日上涨概率(%)', 0),
                results.get('5日上涨概率(%)', 0),
                results.get('10日上涨概率(%)', 0),
                results.get('20日上涨概率(%)', 0)
            ]

            # 添加第一行数据
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=2, column=col_idx, value=value)
                cell.alignment = data_alignment
                cell.border = thin_border

            # 添加第二行数据
            ws.cell(row=3, column=5, value='平均收益率').alignment = data_alignment
            for idx, period in enumerate(StrategyConfig.HOLDING_PERIODS, 6):
                ws.cell(row=3, column=idx, value=results.get('{}日平均收益率(%)'.format(period), 0))

            # 设置单元格样式
            for row in ws.iter_rows(min_row=2, max_row=3):
                for cell in row:
                    cell.alignment = data_alignment
                    cell.alignment = Alignment(wrap_text=True)
                    cell.border = thin_border

            # 合并单元格
            for col_idx in range(1, 5):
                ws.merge_cells(start_row=2, start_column=col_idx, end_row=3, end_column=col_idx)

            # 设置列宽
            col_widths = [24, 48, 24, 24, 24, 24, 24, 24, 24, 24]
            for idx, width in enumerate(col_widths, 1):
                col_letter = get_column_letter(idx)
                ws.column_dimensions[col_letter].width = width

            # 设置行高
            for row_idx in range(1, 4):
                ws.row_dimensions[row_idx].height = 35

            # 添加买入详情工作表
            buy_signals = self.strategy_buy_details.get(strategy_name)
            # 在 generate_report 中增加空结果检查
            if buy_signals is None or buy_signals.empty:
                logger.warning("策略 {} 无有效买入信号".format(strategy_name))
                continue

            # 处理买入详情工作表时使用中文表头
            if buy_signals is not None and not buy_signals.empty:
                # 按 date 降序排序，并保留前100条数据
                buy_signals_sorted = buy_signals.sort_values(by='date', ascending=False).head(100)
                detail_ws = master_wb.create_sheet(title="{}详情".format(safe_name[:28]))

                # 写入表头（使用中文映射）
                columns = buy_signals_sorted.columns.tolist()
                for col_idx, col_name in enumerate(columns, 1):
                    # 使用中文列名，如果没有映射则保留原列名
                    chinese_name = StrategyConfig.DETAIL_COLUMN_MAPPING.get(col_name, col_name)
                    cell = detail_ws.cell(row=1, column=col_idx, value=chinese_name)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = thin_border

                # 写入数据
                for row_idx, row in enumerate(buy_signals_sorted.itertuples(index=False), 2):
                    for col_idx, value in enumerate(row, 1):
                        cell = detail_ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.alignment = data_alignment
                        cell.border = thin_border

                # 设置列宽
                for col_idx in range(1, len(columns) + 1):
                    col_letter = get_column_letter(col_idx)
                    detail_ws.column_dimensions[col_letter].width = 20

        # 保存主报告文件
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        master_file = os.path.join(report_dir, "策略汇总报告_{}.xlsx".format(timestamp))

        # 保存报告并记录文件路径
        try:
            master_wb.save(master_file)
            logger.info("所有策略报告已汇总保存至: {}".format(master_file))
            self.report_file_path = master_file  # 保存报告路径
        except PermissionError:
            alt_file = os.path.join(report_dir, "策略报告_{}.xlsx".format(timestamp))
            master_wb.save(alt_file)
            logger.warning("原文件被占用，报告已保存为: {}".format(alt_file))
            self.report_file_path = alt_file  # 保存替代报告路径
        except Exception as e:
            logger.error(f"保存策略报告出错: {e}")
            self.report_file_path = None
            return False

        # 打印报告摘要
        logger.info("\n策略回测报告摘要:")
        for strategy_name, results in self.strategy_results.items():
            logger.info(f"\n策略: {strategy_name}")
            logger.info(f"买入次数: {results['买入次数']}")
            for period in StrategyConfig.HOLDING_PERIODS:
                win_prob = results.get(f'{period}日上涨概率(%)', 0)
                avg_return = results.get(f'{period}日平均收益率(%)', 0)
                logger.info(f"{period}日上涨概率: {win_prob:.2f}%")
                logger.info(f"{period}日平均收益率: {avg_return:.2f}%")

        return True

    def get_today_matching_stocks(self, strategies):
        """获取当天符合策略的股列表，优化内存占用"""
        if self.return_data is None:
            logger.warning("没有数据用于筛选当日策略股票")
            return {}, None

        # 获取最新交易日（数据中的最后一天）- 使用视图而非副本
        latest_date = self.return_data['date'].max()
        logger.info(f"获取 {latest_date.strftime('%Y-%m-%d')} 符合策略的股票列表")

        today_matches = {}
        # 确定要运行的策略
        all_strategies = StrategyRegistry.list_strategies()
        if 'all' in strategies:
            strategies_to_run = all_strategies
        else:
            strategies_to_run = [s for s in strategies if s in all_strategies]

        logger.info(f"将要运行的策略: {', '.join(strategies_to_run)}")

        # 只保留最新交易日数据，减少后续处理的数据量
        latest_data = self.return_data[self.return_data['date'] == latest_date]

        # 提取所需列，减少内存占用
        required_cols = ['code', 'date', 'close', 'pct_chg']
        if 'name' in self.return_data.columns:
            required_cols.insert(1, 'name')

        # 只保留必要列
        latest_data = latest_data[required_cols]

        for strategy_name in strategies_to_run:
            # 获取策略函数
            strategy_func = StrategyRegistry.get_strategy(strategy_name)
            if not strategy_func:
                logger.warning(f"策略 {strategy_name} 未注册，跳过")
                continue

            try:
                # 直接在原始数据上应用策略，避免复制
                # 策略函数应优化为不修改原始数据
                strategy_data = strategy_func(self.return_data)

                # 只筛选最新交易日的买入信号
                if 'buy_signal' in strategy_data.columns:
                    # 使用布尔索引直接筛选，避免中间变量
                    mask = (strategy_data['date'] == latest_date) & (strategy_data['buy_signal'] == 1)
                    if mask.any():
                        # 合并结果，只获取需要的列
                        result = latest_data[mask[latest_data.index]]
                        today_matches[strategy_name] = result.sort_values('code')
                        logger.info(f"策略 {strategy_name} 当日匹配 {len(today_matches[strategy_name])} 只股票")
                    else:
                        today_matches[strategy_name] = pd.DataFrame(columns=required_cols)
            finally:
                # 及时清理不再使用的大对象
                if 'strategy_data' in locals():
                    del strategy_data
                gc.collect()  # 主动触发垃圾回收

        return today_matches, latest_date

    def export_today_matches_json(self, today_matches, date, output_path):
        """导出当日策略候选股为JSON，供后续融合分析使用"""
        try:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
        except Exception:
            pass

        payload = {
            'trade_date': date.strftime('%Y-%m-%d') if hasattr(date, 'strftime') else str(date),
            'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'strategies': {},
            'codes': [],
        }

        all_codes = set()
        for strategy, df in (today_matches or {}).items():
            if df is None or df.empty:
                payload['strategies'][strategy] = []
                continue

            records = []
            for row in df.to_dict(orient='records'):
                record = {}
                for k, v in row.items():
                    if hasattr(v, 'strftime'):
                        record[k] = v.strftime('%Y-%m-%d')
                    elif pd.isna(v):
                        record[k] = None
                    else:
                        record[k] = v.item() if hasattr(v, 'item') else v
                records.append(record)
                code = str(record.get('code', '')).strip()
                if code:
                    all_codes.add(code)

            payload['strategies'][strategy] = records

        payload['codes'] = sorted(all_codes)
        payload['total_codes'] = len(payload['codes'])

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

        logger.info(f"已导出候选股JSON: {output_path}，共 {payload['total_codes']} 只股票")
        return output_path

    def send_strategy_email(self, today_matches, date, enhancer_json_path=None):
        """发送包含当日策略匹配股票的邮件，附带Excel报告作为附件，顶部增加候选增强分析结果"""
        if not today_matches:
            logger.warning("没有匹配的股票数据可发送邮件")
            return False

        # 格式化日期字符串
        date_str = date.strftime('%Y-%m-%d')

        # 创建邮件
        msg = MIMEMultipart()
        msg['From'] = EmailConfig.SMTP_USER
        msg['To'] = ", ".join(EmailConfig.RECIPIENTS)
        msg['Subject'] = f"{date_str}-report"

        # 邮件正文构建：顶部先放候选增强分析结果（如果有）
        body_parts = []

        # === 新增：候选增强分析结果放在最上方 ===
        if enhancer_json_path and os.path.exists(enhancer_json_path):
            try:
                import json
                with open(enhancer_json_path, 'r', encoding='utf-8') as f:
                    enhancer_result = json.load(f)
                top_candidates = enhancer_result.get('top_candidates', [])
                if top_candidates:
                    body_parts.append(f"<h2>🔝 {date_str} 候选增强分析 - 最可能盈利 Top {len(top_candidates)}</h2>")
                    body_parts.append("<table border='1' cellpadding='4' cellspacing='0' style='border-collapse:collapse;'>")
                    body_parts.append("<thead><tr>")
                    body_parts.append("<th>排名</th><th>代码</th><th>名称</th><th>当前价</th><th>涨跌幅</th><th>最终得分</th><th>盈利概率</th><th>动作</th><th>总结</th>")
                    body_parts.append("</tr></thead><tbody>")
                    for idx, cand in enumerate(top_candidates, 1):
                        if cand.get('status') != 'ok':
                            continue
                        code = cand.get('code', '')
                        name = cand.get('name', '')
                        price = cand.get('price', '')
                        change = cand.get('change_pct', '')
                        score = cand.get('final_score', '')
                        prob = cand.get('profit_probability', 0) * 100
                        action = cand.get('action', '')
                        summary = cand.get('summary', '')
                        body_parts.append(f"<tr><td>{idx}</td><td>{code}</td><td>{name}</td><td>{price:.2f}</td><td>{change:.2f}%</td><td>{score:.1f}</td><td>{prob:.0f}%</td><td>{action}</td><td>{summary}</td></tr>")
                    body_parts.append("</tbody></table><br><hr><br>")
                    logger.info(f"已将候选增强 Top {len(top_candidates)} 添加到邮件正文顶部")
            except Exception as e:
                logger.error(f"加载/添加候选增强结果失败: {str(e)}，跳过该部分")

        # 原有正文：策略匹配列表
        body_parts.append(f"<h3>{date.strftime('%Y-%m-%d')} 策略匹配列表</h3>")

        for strategy, df in today_matches.items():
            # 获取策略描述
            strategy_desc = StrategyRegistry.get_description(strategy)

            # 添加策略标题和描述
            body_parts.append(f"<h4>{strategy}（共 {len(df)} 只）</h4>")
            body_parts.append(f"<p style='font-size:12px; color:#666;'>{strategy_desc}</p>")

            if not df.empty:
                # 转换列名为中文
                chinese_columns = {col: StrategyConfig.DETAIL_COLUMN_MAPPING.get(col, col) for col in df.columns}
                # 直接使用rename的inplace参数减少复制
                df_renamed = df.rename(columns=chinese_columns)
                # 转换为HTML表格
                body_parts.append(df_renamed.to_html(index=False, justify='left', classes='table table-striped'))
            else:
                body_parts.append("<p>无匹配股票</p>")

            body_parts.append("<br>")

            # 及时清理不再需要的DataFrame
            del df

        # 合并所有部分，减少字符串拼接次数
        body = ''.join(body_parts)
        msg.attach(MIMEText(body, 'html', 'utf-8'))

        # 清理不再需要的变量
        del body_parts, body

        # 添加Excel报告作为附件 - 仅在需要时读取
        if self.report_file_path and os.path.exists(self.report_file_path):
            try:
                # 使用上下文管理器读取文件，自动关闭
                with open(self.report_file_path, "rb") as attachment:
                    part = MIMEBase("application", "octet-stream")
                    part.set_payload(attachment.read())

                encoders.encode_base64(part)
                filename = f"{date_str}-report.xlsx"
                part.add_header(
                    "Content-Disposition",
                    "attachment",
                    filename = ("utf-8", "", filename)
                )
                msg.attach(part)
                logger.info(f"已添加附件: {filename}")
            except Exception as e:
                logger.error(f"添加邮件附件失败: {str(e)}")
            finally:
                gc.collect()

        # 尝试发送邮件，添加重试机制
        max_retries = 5
        for attempt in range(max_retries):
            server = None
            try:
                logger.info(f"尝试使用SSL连接... (尝试 {attempt + 1}/{max_retries})")
                context = ssl.create_default_context()
                server = smtplib.SMTP_SSL(EmailConfig.SMTP_SERVER, EmailConfig.SMTP_PORT, context=context)
                logger.info("已连接到SMTP服务器，准备登录...")
                server.login(EmailConfig.SMTP_USER, EmailConfig.SMTP_PASSWORD)
                logger.info("登录成功，准备发送邮件...")
                text = msg.as_string()
                server.sendmail(EmailConfig.SMTP_USER, EmailConfig.RECIPIENTS, text)
                logger.info("策略匹配股票邮件发送成功（含附件，顶部包含候选增强结果）")
                return True
            except Exception as e:
                logger.error(f"发送邮件失败 (尝试 {attempt + 1}/{max_retries}): {str(e)}")
                if attempt < max_retries - 1:
                    logger.info("等待5秒后重试...")
                    time.sleep(5)
            finally:
                if server:
                    try:
                        server.quit()
                    except Exception as e:
                        # 忽略连接关闭时的小错误
                        logger.debug(f"关闭SMTP连接时出现小问题（可忽略）: {str(e)}")

        return False


# ----------------------------
# 交易策略实现（使用注册器）
# ----------------------------
@StrategyRegistry.register("均线交叉策略")
def optimized_ma_crossover_strategy(df):
    """
    优化的移动平均线交叉策略
    结合成交量、MACD和RSI指标过滤信号
    """

    def _calculate_macd(group):
        """计算MACD指标"""
        close = group['close']
        # 计算EMA
        ema_fast = close.ewm(span=StrategyConfig.MACD_FAST, adjust=False).mean()
        ema_slow = close.ewm(span=StrategyConfig.MACD_SLOW, adjust=False).mean()

        # 计算DIF和DEA
        dif = ema_fast - ema_slow
        dea = dif.ewm(span=StrategyConfig.MACD_SIGNAL, adjust=False).mean()

        # 计算MACD柱
        macd_hist = 2 * (dif - dea)

        return dif, dea, macd_hist

    def _calculate_rsi(group):
        """计算RSI指标"""
        close = group['close']
        delta = close.diff()

        # 计算上涨和下跌的幅度
        gain = delta.where(delta > 0, 0)
        loss = -delta.where(delta < 0, 0)

        # 计算平均收益和平均损失
        avg_gain = gain.rolling(window=StrategyConfig.RSI_WINDOW, min_periods=1).mean()
        avg_loss = loss.rolling(window=StrategyConfig.RSI_WINDOW, min_periods=1).mean()

        # 计算相对强度(RS)和RSI
        rs = avg_gain / avg_loss
        rsi = 100 - (100 / (1 + rs))

        return rsi

    def _apply_group(group):
        group = group.sort_values('date')

        # 1. 计算均线
        group['ma5'] = group['close'].rolling(window=StrategyConfig.MA_WINDOWS['short']).mean()
        group['ma20'] = group['close'].rolling(window=StrategyConfig.MA_WINDOWS['long']).mean()
        # 新增60日均线
        group['ma60'] = group['close'].rolling(window=60).mean()

        # 2. 计算成交量指标
        group['vol_ma5'] = group['volume'].rolling(window=StrategyConfig.VOL_WINDOW).mean()
        # 计算连续2日成交量大于5日均量
        group['vol_above_ma5'] = group['volume'] > group['vol_ma5']
        group['vol_above_ma5_2d'] = group['vol_above_ma5'].rolling(2).sum() == 2

        # 3. 计算MACD指标
        dif, dea, macd_hist = _calculate_macd(group)
        group['macd_dif'] = dif
        group['macd_dea'] = dea
        group['macd_hist'] = macd_hist

        # 4. 计算RSI指标
        group['rsi'] = _calculate_rsi(group)

        # 5. 初始化信号列
        group['buy_signal'] = 0
        group['sell_signal'] = 0

        # 6. 生成基础买入信号（5日均线上穿20日均线）
        cross_up = (group['ma5'] > group['ma20']) & (group['ma5'].shift(1) <= group['ma20'].shift(1))

        # 7. 生成基础卖出信号（5日均线下穿20日均线）
        cross_down = (group['ma5'] < group['ma20']) & (group['ma5'].shift(1) >= group['ma20'].shift(1))

        # 8. 高级信号过滤条件
        # 买入信号增强条件（新增三个条件）
        buy_conditions = (
                cross_up &  # 金叉条件
                (group['volume'] > 1.5 * group['vol_ma5']) &  # 成交量放大50%以上
                (group['macd_hist'] > 0) &  # MACD柱状线为正（多头趋势）
                (group['macd_dif'] > group['macd_dea']) &  # DIF在DEA上方
                (group['rsi'] > StrategyConfig.RSI_OVERSOLD) &  # RSI脱离超卖区
                (group['rsi'] < StrategyConfig.RSI_OVERBOUGHT) &  # RSI未超买
                (group['close'] > group['open']) &  # 新增：当日阳线（收盘价>开盘价）
                group['vol_above_ma5_2d'] &  # 新增：连续2日成交量>5日均量
                (group['close'] > group['ma60'])  # 新增：收盘价>60日均价
        )

        # 卖出信号增强条件
        sell_conditions = (
                cross_down &  # 死叉条件
                (group['volume'] > group['vol_ma5']) &  # 成交量放大
                (group['macd_hist'] < 0) &  # MACD柱状线为负（空头趋势）
                (group['macd_dif'] < group['macd_dea']) &  # DIF在DEA下方
                (group['rsi'] > StrategyConfig.RSI_OVERBOUGHT)  # RSI超买
        )

        # 9. 应用高级信号过滤
        group.loc[buy_conditions, 'buy_signal'] = 1
        group.loc[sell_conditions, 'sell_signal'] = 1

        # 10. 添加趋势强度指标
        group['trend_strength'] = (group['ma5'] - group['ma20']) / group['ma20'] * 100

        return group

    return df.groupby(df['code'], group_keys=False, as_index=False).apply(_apply_group)

@StrategyRegistry.register("成交量突破策略")
def volume_breakout_strategy(df):
    """成交量突破策略"""

    def _apply_group(group):
        group = group.sort_values('date')
        group['buy_signal'] = 0
        # 计算成交量均值和标准差
        group['vol_mean'] = group['volume'].rolling(window=StrategyConfig.MA_WINDOWS['long']).mean()
        group['vol_std'] = group['volume'].rolling(window=StrategyConfig.MA_WINDOWS['long']).std()
        # 生成买入信号
        breakout_cond = group['volume'] > group['vol_mean'] + StrategyConfig.VOLUME_THRESHOLDS['breakout'] * group[
            'vol_std']
        group.loc[breakout_cond, 'buy_signal'] = 1
        return group

    return df.groupby(df['code'], group_keys=False, as_index=False).apply(_apply_group)


@StrategyRegistry.register("多均线共振策略",
                           "6均线齐头向上且顺序排列(5>10>20>60>90>120)，5日金叉10日，成交量放大，MACD(14,53,5)金叉且量柱翻红，布林带刚开口，RSI三线向上")
def multi_ma_resonance_strategy(df):
    """
    多均线共振策略（优化6均线顺序排列条件）
    1、5、10、20、60、90、120日6条均线齐头向上，且按5>10>20>60>90>120顺序排列无交叉，5日均线刚好金叉10日均线
    2、成交量放大
    3、MACD（14,53,5）金叉，且量柱由绿翻红
    4、boll带刚开口（从收缩转为扩张的瞬间）
    5、RSI三线齐头向上
    """

    def _calculate_macd(group):
        """计算MACD指标(14,53,5)"""
        close = group['close']
        ema_fast = close.ewm(span=14, adjust=False).mean()
        ema_slow = close.ewm(span=53, adjust=False).mean()
        dif = ema_fast - ema_slow
        dea = dif.ewm(span=5, adjust=False).mean()
        macd_hist = 2 * (dif - dea)
        return dif, dea, macd_hist

    def _calculate_bollinger(group, window=20, num_std=2):
        """计算布林带"""
        close = group['close']
        ma = close.rolling(window=window).mean()
        std = close.rolling(window=window).std()
        upper = ma + num_std * std
        lower = ma - num_std * std
        return upper, lower, ma

    def _calculate_rsi_lines(group):
        """计算RSI三线(6,12,24)"""
        close = group['close']
        delta = close.diff()

        gain = delta.where(delta > 0, 0)
        loss = -delta.where(delta < 0, 0)

        rsi6 = 100 - (100 / (1 + gain.rolling(6).mean() / loss.rolling(6).mean()))
        rsi12 = 100 - (100 / (1 + gain.rolling(12).mean() / loss.rolling(12).mean()))
        rsi24 = 100 - (100 / (1 + gain.rolling(24).mean() / loss.rolling(24).mean()))

        return rsi6, rsi12, rsi24

    def _apply_group(group):
        group = group.sort_values('date')

        # 1. 计算6条均线系统
        group['ma5'] = group['close'].rolling(window=5).mean()
        group['ma10'] = group['close'].rolling(window=10).mean()
        group['ma20'] = group['close'].rolling(window=20).mean()
        group['ma60'] = group['close'].rolling(window=60).mean()
        group['ma90'] = group['close'].rolling(window=90).mean()
        group['ma120'] = group['close'].rolling(window=120).mean()

        # 2. 计算成交量指标
        group['vol_ma20'] = group['volume'].rolling(window=20).mean()
        # group['vol_increase'] = group['volume'] > StrategyConfig.VOLUME_THRESHOLDS['resonance'] * group['vol_ma20']
        group['vol_increase'] = group['volume'] > group['vol_ma20']

        # 3. 计算MACD指标
        dif, dea, macd_hist = _calculate_macd(group)
        group['macd_dif'] = dif
        group['macd_dea'] = dea
        group['macd_hist'] = macd_hist

        # 4. 计算布林带及刚开口条件
        upper, lower, boll_ma = _calculate_bollinger(group)
        group['boll_upper'] = upper
        group['boll_lower'] = lower
        group['boll_width'] = upper - lower
        group['boll_just_opened'] = (
                (group['boll_width'] > group['boll_width'].shift(1)) &  # 今日比昨日宽
                (group['boll_width'].shift(1) < group['boll_width'].shift(2))  # 昨日比前日窄
        )

        # 5. 计算RSI三线
        rsi6, rsi12, rsi24 = _calculate_rsi_lines(group)
        group['rsi6'] = rsi6
        group['rsi12'] = rsi12
        group['rsi24'] = rsi24

        # 6. 构建买入条件
        # 均线条件：6条均线齐头向上+顺序排列（5>10>20>60>90>120）+5日金叉10日
        ma_cond = (
            # 所有均线向上（当前值>前一天值）
                (group['ma5'] > group['ma5'].shift(1)) &
                (group['ma10'] > group['ma10'].shift(1)) &
                (group['ma20'] > group['ma20'].shift(1)) &
                (group['ma60'] > group['ma60'].shift(1)) &
                (group['ma90'] > group['ma90'].shift(1)) &
                (group['ma120'] > group['ma120'].shift(1)) &
                # 均线顺序排列（从上到下），无交叉
                (group['ma5'] > group['ma10']) &
                (group['ma10'] > group['ma20']) &
                (group['ma20'] > group['ma60']) &
                (group['ma60'] > group['ma90']) &
                (group['ma90'] > group['ma120']) &
                # 5日刚金叉10日
                (group['ma5'] > group['ma10']) &
                (group['ma5'].shift(1) <= group['ma10'].shift(1))
        )

        # MACD条件：金叉且量柱由绿翻红
        macd_cond = (
                (group['macd_dif'] > group['macd_dea']) &
                (group['macd_dif'].shift(1) <= group['macd_dea'].shift(1)) &
                (group['macd_hist'] > 0) &
                (group['macd_hist'].shift(1) <= 0)
        )

        # RSI条件：三线齐头向上且未超买
        rsi_cond = (
                (group['rsi6'] > group['rsi6'].shift(1)) &
                (group['rsi12'] > group['rsi12'].shift(1)) &
                (group['rsi24'] > group['rsi24'].shift(1)) &
                (group['rsi6'] < StrategyConfig.RSI_OVERBOUGHT) &
                (group['rsi12'] < StrategyConfig.RSI_OVERBOUGHT) &
                (group['rsi24'] < StrategyConfig.RSI_OVERBOUGHT)
        )

        # 综合买入信号
        group['buy_signal'] = 0
        group.loc[
            ma_cond &
            group['vol_increase'] &
            macd_cond &
            group['boll_just_opened'] &
            rsi_cond,
            'buy_signal'
        ] = 1

        return group

    # 按股票代码分组应用策略
    return df.groupby('code', group_keys=False).apply(_apply_group)


@StrategyRegistry.register("神奇九转策略")
def optimized_wonderful_9_turn_strategy(df):
    """
    优化的神奇九转策略
    解决多进程pickle问题，将内部函数改为外部函数
    """
    # 并行处理优化
    if len(df) > 10000:  # 大数据集使用并行处理
        import multiprocessing as mp
        groups = [group for _, group in df.groupby('code')]
        with mp.Pool(processes=min(mp.cpu_count(), 8)) as pool:
            results = pool.map(_wonderful_9_turn_apply_group, groups)
        result = pd.concat(results, ignore_index=True)
    else:
        result = df.groupby('code', group_keys=False, observed=True).apply(_wonderful_9_turn_apply_group)

    return result

def _wonderful_9_turn_apply_group(group):
    """
    优化版神奇九转策略
    主要改进：条件权重化、参数可配置化、计算效率优化
    """

    def _calculate_rsi(close, window=StrategyConfig.RSI_WINDOW):
        """优化版RSI计算"""
        delta = close.diff()
        gain = delta.where(delta > 0, 0)
        loss = -delta.where(delta < 0, 0)

        avg_gain = gain.rolling(window=window, min_periods=1).mean()
        avg_loss = loss.rolling(window=window, min_periods=1).mean()

        rs = avg_gain / avg_loss.replace(0, 0.001)
        return 100 - (100 / (1 + rs))

    def _calculate_macd(close):
        """优化版MACD计算"""
        ema_fast = close.ewm(span=StrategyConfig.MACD_FAST, adjust=False).mean()
        ema_slow = close.ewm(span=StrategyConfig.MACD_SLOW, adjust=False).mean()

        dif = ema_fast - ema_slow
        dea = dif.ewm(span=StrategyConfig.MACD_SIGNAL, adjust=False).mean()
        macd_hist = 2 * (dif - dea)

        return macd_hist


    group = group.sort_values('date').copy()

    # 1. 基础神奇九转条件
    close_lt_4days = group['close'] < group['close'].shift(4)
    streak_count = close_lt_4days.rolling(window=9).sum()

    # 低点比较条件
    low_8_lt_6 = group['low'].shift(1) < group['low'].shift(3)
    low_8_lt_7 = group['low'].shift(1) < group['low'].shift(2)
    low_9_lt_6 = group['low'] < group['low'].shift(3)
    low_9_lt_7 = group['low'] < group['low'].shift(2)
    low_condition = low_8_lt_6 | low_8_lt_7 | low_9_lt_6 | low_9_lt_7

    base_signal = (streak_count == 9) & low_condition

    # 2. 技术指标过滤（使用权重系统）
    conditions_met = 0
    total_conditions = 5

    # RSI超卖条件（权重：1）
    group['rsi'] = _calculate_rsi(group['close'])
    rsi_condition = group['rsi'] < getattr(StrategyConfig, 'RSI_OVERSOLD', 40)
    conditions_met += rsi_condition.astype(int)

    # MACD动量条件（权重：1）
    group['macd_hist'] = _calculate_macd(group['close'])
    macd_improving = (group['macd_hist'] > group['macd_hist'].shift(1))
    macd_negative = group['macd_hist'] < 0
    macd_condition = macd_improving & macd_negative
    conditions_met += macd_condition.astype(int)

    # 成交量条件（权重：1）
    group['vol_ma20'] = group['volume'].rolling(window=20).mean()
    volume_condition = group['volume'] > StrategyConfig.VOLUME_THRESHOLDS['resonance'] * group['vol_ma20']
    conditions_met += volume_condition.astype(int)

    # 趋势条件（权重：1）
    group['ma20'] = group['close'].rolling(window=20).mean()
    group['ma60'] = group['close'].rolling(window=60).mean()
    trend_up = (group['ma20'] > group['ma60']) & \
               (group['ma20'] > group['ma20'].shift(1)) & \
               (group['ma60'] > group['ma60'].shift(1))
    conditions_met += trend_up.astype(int)

    # 均线支撑条件（权重：1）
    price_near_ma = group['close'] >= getattr(StrategyConfig, 'MA_SUPPORT_RATIO', 0.98) * group['ma20']
    conditions_met += price_near_ma.astype(int)

    # 3. 动态阈值系统
    required_conditions = getattr(StrategyConfig, 'MIN_CONDITIONS', 3)
    filter_signal = conditions_met >= required_conditions

    # 4. 生成最终信号
    group['buy_signal'] = 0
    final_signal = base_signal & filter_signal

    # 信号确认：收盘价确认
    final_signal = final_signal & (group['close'] > group['open'])

    group.loc[final_signal, 'buy_signal'] = 1

    # 5. 信号强度评分（用于仓位管理）
    group['signal_strength'] = conditions_met / total_conditions

    return group

@StrategyRegistry.register("N字反包策略")
def n_shape_reversal_strategy(df):
    """N字反包交易策略"""

    def _apply_group(group):
        group = group.sort_values('date')
        group['buy_signal'] = 0
        # 向量化判断N字形态
        prev_high = group['high'].shift(2)
        callback_close = group['close'].shift(1)
        current_close = group['close']
        mask = (callback_close < prev_high) & (current_close > prev_high)
        group.loc[mask, 'buy_signal'] = 1
        return group

    return df.groupby(df['code'], group_keys=False, as_index=False).apply(_apply_group)


@StrategyRegistry.register("涨停回调策略")
def limit_up_pullback_strategy(df):
    """
    涨停回调策略（优化版）
    满足条件：
    1、90天内首次涨停板（核心前提）
    2、涨停时突破90天内新高
    3、之后缩量回调到涨停前最高位附近，最低价格接受破位2％
    4、回调至涨停前最高位附近至少两次后，出现放量上涨时买入
    """

    def _apply_group(group):
        # 按交易日期排序并复制，避免SettingWithCopyWarning
        group = group.sort_values('date').copy()
        group['buy_signal'] = 0

        # 1. 定义涨停条件 (涨幅≥10%)
        limit_up_cond = (group['pct_chg'] >= StrategyConfig.LIMIT_UP_PCT_THRESHOLD)
        group['is_limit_up'] = limit_up_cond

        # 标记90天内首次涨停（核心前提条件）
        # 计算过去90天内是否有过涨停
        group['has_limit_up_90d'] = group['is_limit_up'].rolling(window=90, min_periods=1).sum().shift(1,
                                                                                                       fill_value=0) > 0
        # 首次涨停：当天是涨停，且过去90天内没有涨停
        group['first_limit_up_90d'] = limit_up_cond & (~group['has_limit_up_90d'])

        # 2. 标记涨停时突破90天内新高的情况
        group['high_90d'] = group['high'].rolling(window=90, min_periods=1).max()
        # 有效涨停：首次涨停且突破90天新高
        group['valid_limit_up'] = group['first_limit_up_90d'] & (group['high'] >= group['high_90d'])

        # 3. 记录有效涨停的关键价格和信息
        # 涨停价、涨停前最高价、涨停日成交量
        group['limit_up_price'] = np.nan
        group['pre_limit_high'] = np.nan
        group['limit_up_volume'] = np.nan
        group['limit_up_date'] = pd.NaT  # 使用pd.NaT而不是np.nan来表示缺失的日期时间

        # 只在有效涨停时更新上述信息
        valid_mask = group['valid_limit_up']
        if not valid_mask.empty:
            # 直接赋值 datetime 类型，无需转换为 values
            group.loc[valid_mask, 'limit_up_date'] = pd.to_datetime(group.loc[valid_mask, 'date'])

        group.loc[valid_mask, 'limit_up_price'] = group.loc[valid_mask, 'high']
        group.loc[valid_mask, 'limit_up_volume'] = group.loc[valid_mask, 'volume']
        # 使用date列而不是索引，避免数据类型问题
        group.loc[valid_mask, 'limit_up_date'] = group.loc[valid_mask, 'date']
        # 涨停前的最高价（取涨停前89天内的最高价）
        group.loc[valid_mask, 'pre_limit_high'] = group.loc[valid_mask, 'high'].shift(1).rolling(window=89,
                                                                                                 min_periods=1).max()

        group['limit_up_price'] = group['limit_up_price'].ffill()
        group['pre_limit_high'] = group['pre_limit_high'].ffill()
        group['limit_up_volume'] = group['limit_up_volume'].ffill()
        group['limit_up_date'] = group['limit_up_date'].ffill().infer_objects(copy=False)

        # 4. 确定回调区间 (涨停前最高价上下2%)
        group['pullback_low'] = group['pre_limit_high'] * 0.98  # 允许2%破位
        group['pullback_high'] = group['pre_limit_high'] * 1.02

        # 5. 缩量条件 (成交量低于涨停日成交量的50%)
        group['is_shrinking_volume'] = group['volume'] < group['limit_up_volume'] * 0.5

        # 6. 标记处于回调区间内（仅在有有效涨停记录时才判断）
        has_valid_limit_up = ~group['limit_up_price'].isna()
        in_pullback_range = (group['low'] >= group['pullback_low']) & (group['high'] <= group['pullback_high'])
        group['in_pullback_zone'] = has_valid_limit_up & in_pullback_range & group['is_shrinking_volume']

        # 7. 计数回调至区间的次数（仅在有效涨停后计数）
        group['pullback_count'] = 0
        pullback_counter = 0
        last_limit_up_idx = None

        for i in range(len(group)):
            # 遇到新的有效涨停，重置计数器
            if group['valid_limit_up'].iloc[i]:
                pullback_counter = 0
                last_limit_up_idx = i
                continue

            # 如果没有有效涨停记录，不计数
            if last_limit_up_idx is None:
                continue

            # 只对有效涨停后的交易日进行计数
            if i <= last_limit_up_idx:
                continue

            # 进入回调区间且缩量，计数+1（跨天连续回调只算一次）
            current_in = group['in_pullback_zone'].iloc[i]
            prev_in = group['in_pullback_zone'].iloc[i - 1] if i > 0 else False

            if current_in and not prev_in:
                pullback_counter += 1

            group.iloc[i, group.columns.get_loc('pullback_count')] = pullback_counter

        # 8. 放量上涨条件 (成交量大于5日均量1.5倍，且收盘价上涨)
        group['vol_ma5'] = group['volume'].rolling(window=5, min_periods=1).mean()
        group['is_rising_volume'] = (group['volume'] > group['vol_ma5'] * 1.5) & (group['pct_chg'] > 0)

        # 9. 最终买入信号：所有条件同时满足
        buy_cond = (
                has_valid_limit_up &  # 存在有效涨停
                (group['pullback_count'] >= 2) &  # 至少回调2次
                group['is_rising_volume']  # 放量上涨
        )

        group.loc[buy_cond, 'buy_signal'] = 1

        # 清理临时列，减少内存占用
        temp_cols = ['is_limit_up', 'has_limit_up_90d', 'high_90d', 'valid_limit_up',
                     'limit_up_price', 'limit_up_volume', 'limit_up_date', 'pre_limit_high',
                     'pullback_low', 'pullback_high', 'is_shrinking_volume',
                     'in_pullback_zone', 'vol_ma5', 'is_rising_volume']
        group = group.drop(columns=[col for col in temp_cols if col in group.columns])

        return group

    # 按股票代码分组处理
    return df.groupby('code', group_keys=False, as_index=False).apply(_apply_group)


@StrategyRegistry.register("连续平稳后涨停策略")
def continuous_stable_then_limit_up_strategy(df):
    """连续平稳后涨停策略"""

    def _apply_group(group):
        group = group.sort_values('date')
        group['buy_signal'] = 0
        # 连续10日涨跌幅在阈值以内
        stable_mask = group['pct_chg'].abs().rolling(window=10).apply(
            lambda x: all(x <= StrategyConfig.STABILITY_THRESHOLD), raw=True
        )
        # 第11日涨停
        mask = (stable_mask.shift(1) == 1) & (group['pct_chg'] >= StrategyConfig.LIMIT_UP_PCT_THRESHOLD)
        group.loc[mask, 'buy_signal'] = 1
        return group

    return df.groupby(df['code'], group_keys=False, as_index=False).apply(_apply_group)


@StrategyRegistry.register("价格突破策略")
def price_breakout_strategy(df):
    """价格突破策略"""

    def _apply_group(group):
        group = group.sort_values('date')
        group['buy_signal'] = 0
        group['high_20d'] = group['high'].rolling(window=StrategyConfig.MA_WINDOWS['long']).max()
        # 收盘价突破20日最高价
        group.loc[group['close'] > group['high_20d'].shift(1), 'buy_signal'] = 1
        return group

    return df.groupby(df['code'], group_keys=False, as_index=False).apply(_apply_group)


@StrategyRegistry.register("MACD月线金叉+20日线策略")
def macd_ma_volume_strategy(df):
    """MACD月线金叉+20日线策略"""

    def _apply_group(group):
        group = group.sort_values('date')
        group['buy_signal'] = 0
        group['date'] = pd.to_datetime(group['date'])

        # 标记涨停日
        group['limit_up'] = (group['pct_chg'] >= StrategyConfig.LIMIT_UP_PCT_THRESHOLD).astype(int)
        # 13日内涨停次数
        group['limit_up_13d'] = group['limit_up'].rolling(13, min_periods=1).sum().shift(1)

        # 排除3连板及以上
        group['consecutive_limit'] = group['limit_up'].groupby(
            (group['limit_up'].diff() != 0).cumsum()
        ).cumcount() + 1
        invalid_mask = group['consecutive_limit'] >= 3

        # 月线MACD计算
        monthly = group.set_index('date').resample('ME')['close'].last().to_frame('close').reset_index()
        if len(monthly) >= 26:
            monthly['DIF'] = monthly['close'].ewm(span=12, adjust=False).mean() - monthly['close'].ewm(span=26,
                                                                                                       adjust=False).mean()
            monthly['DEA'] = monthly['DIF'].ewm(span=9, adjust=False).mean()
            monthly['macd_golden'] = (monthly['DIF'] > monthly['DEA']) & (
                        monthly['DIF'].shift() <= monthly['DEA'].shift())
        else:
            monthly['macd_golden'] = False

        # 合并月线信号到日线
        group = pd.merge_asof(
            group.sort_values('date'),
            monthly[['date', 'macd_golden']].sort_values('date'),
            on='date',
            direction='backward'
        )

        # 日线指标
        group['ma20'] = group['close'].rolling(StrategyConfig.MA_WINDOWS['long'], min_periods=1).mean()
        group['prev_20d_high'] = group['high'].shift(1).rolling(StrategyConfig.MA_WINDOWS['long'], min_periods=1).max()

        # 综合条件
        vol_mean = group['volume'].shift(1).rolling(StrategyConfig.MA_WINDOWS['long']).mean()
        cond_macd = group['macd_golden']
        cond_limit = group['limit_up_13d'] >= 1
        cond_exclude = ~invalid_mask
        cond_ma = (group['close'] - group['ma20']).abs() / group['ma20'] <= 0.02
        cond_vol = group['volume'] > vol_mean * StrategyConfig.VOLUME_THRESHOLDS['resonance']
        cond_high = (group['close'] >= group['prev_20d_high'] * 0.95) | (group['limit_up'] == 1)

        group.loc[cond_macd & cond_limit & cond_exclude & cond_ma & cond_vol & cond_high, 'buy_signal'] = 1
        return group

    return df.groupby(df['code'], group_keys=False, as_index=False).apply(_apply_group)


@StrategyRegistry.register("低位涨停换手率策略")
def low_position_limit_up_strategy(df):
    """低位涨停换手率策略"""

    def _apply_group(group):
        group = group.sort_values('date')
        group['buy_signal'] = 0

        # 计算60日最高价和最低价
        group['high_60d'] = group['high'].rolling(window=60).max()
        group['low_60d'] = group['low'].rolling(window=60).min()

        # 计算当前价格在60日价格区间的位置
        group['price_position'] = (group['close'] - group['low_60d']) / (group['high_60d'] - group['low_60d']) * 100

        # 低位判断：价格位置低于40%且涨停
        low_position = group['price_position'] < 40
        limit_up = group['pct_chg'] >= StrategyConfig.LIMIT_UP_PCT_THRESHOLD

        # 换手率放大
        group['turnover_ma10'] = group['turn'].rolling(window=10).mean()
        high_turnover = group['turn'] > 1.5 * group['turnover_ma10']

        # 买入信号
        group.loc[low_position & limit_up & high_turnover, 'buy_signal'] = 1
        return group

    return df.groupby(df['code'], group_keys=False, as_index=False).apply(_apply_group)

@StrategyRegistry.register("涨停回调量价共振策略")
def limit_up_callback_resonance_strategy(df):
    """涨停回调量价共振策略"""

    def _apply_group(group):
        group = group.sort_values('date')
        group['buy_signal'] = 0

        # 标记涨停
        group['limit_up'] = (group['pct_chg'] >= StrategyConfig.LIMIT_UP_PCT_THRESHOLD).astype(int)

        # 涨停后5日内回调
        # 修复：处理NaN值，使用fillna(0)填充NaN值后再转换为整数
        group['after_limit_up'] = group['limit_up'].where(group['limit_up'] == 1).ffill().fillna(0).astype(int)
        within_5d = group['after_limit_up'] & (group['after_limit_up'].groupby(
            (group['limit_up'] == 1).cumsum()
        ).cumcount() <= 5)

        # 回调期间缩量
        group['vol_ma5'] = group['volume'].rolling(window=5).mean()
        low_volume = group['volume'] < 0.8 * group['vol_ma5']

        # 20日均线向上
        group['ma20'] = group['close'].rolling(window=20).mean()
        ma20_up = group['ma20'] > group['ma20'].shift(1)

        # 出现阳包阴或曙光初现形态
        bullish_candle = (group['close'] > group['open']) & (group['close'] > group['open'].shift(1))
        morning_star = (group['close'] > (group['open'].shift(1) + group['close'].shift(1)) / 2) & \
                       (group['open'] < group['close'].shift(1)) & \
                       (group['close'].shift(1) < group['open'].shift(1))

        # 买入信号
        buy_mask = within_5d & low_volume & ma20_up & (bullish_candle | morning_star)
        group.loc[buy_mask, 'buy_signal'] = 1
        return group

    return df.groupby(df['code'], group_keys=False, as_index=False).apply(_apply_group)
@StrategyRegistry.register("孕阳线策略")
def big_red_after_black_strategy(df):
    """孕阳线策略"""

    def _apply_group(group):
        group = group.sort_values('date')
        group['buy_signal'] = 0

        # 计算大阴线条件：跌幅≥7%且实体占比≥70%
        prev_pct = group['pct_chg'].shift(1)
        prev_open = group['open'].shift(1)
        prev_close = group['close'].shift(1)
        body_pct = abs(prev_close - prev_open) / prev_open * 100  # 实体占比
        big_black = (prev_pct <= -7) & (body_pct >= 7)

        # 当日阳线且收盘价在前一日开盘价与收盘价之间
        current_bullish = group['close'] > group['open']
        # 当日开盘价高于前一日收盘价至少2%
        high_enough = (group['open'] >= prev_close * 1.02)
        price_in_range = (group['open'] > prev_close) & (group['close'] < prev_open) & high_enough

        # 买入信号
        group.loc[big_black & current_bullish & price_in_range, 'buy_signal'] = 1
        return group

    return df.groupby(df['code'], group_keys=False, as_index=False).apply(_apply_group)

# ----------------------------
# 主函数
# ----------------------------
def main():
    parser = argparse.ArgumentParser(description='股票策略回测系统')
    parser.add_argument('--parallel', action='store_true', help='是否启用并行处理')
    parser.add_argument('--pool_size', type=int, default=2, help='数据库连接池大小')
    parser.add_argument('--start_date', type=str, required=True, help='回测开始日期 (YYYY-MM-DD)')
    parser.add_argument('--end_date', type=str, help='回测结束日期 (YYYY-MM-DD), 默认为今天')
    parser.add_argument('--strategies', nargs='+', default=['all'], help='要回测的策略列表，默认全部')
    parser.add_argument('--send_email', action='store_true', help='是否发送邮件报告')
    parser.add_argument('--export_candidates_json', type=str, help='导出当日策略候选股JSON文件路径')
    parser.add_argument('--force', action='store_true', help='强制执行，忽略非交易日限制')

    args = parser.parse_args()

    # 判断是否为交易日
    if not args.force and not is_trading_day(datetime.now()):
        logger.error("非交易日，程序退出（使用 --force 可强制运行）")
        return

    # 数据库配置
    db_params = {
        'dbname': DbConfig.DB_NAME,
        'user': DbConfig.DB_USER,
        'password': DbConfig.DB_PASSWORD,
        'host': DbConfig.DB_HOST,
        'port': DbConfig.DB_PORT
    }

    # 初始化回测器
    backtest = StockStrategyBacktest(db_params, pool_size=args.pool_size)

    try:
        # 获取股票数据
        end_date = args.end_date or datetime.now().strftime('%Y-%m-%d')
        stock_data = backtest.fetch_stock_data(args.start_date, end_date)
        if stock_data is None:
            logger.error("无法获取股票数据，程序退出")
            return

        # 计算收益率
        return_data = backtest.calculate_returns()
        if return_data is None:
            logger.error("无法计算收益率，程序退出")
            return

        # 确定要运行的策略
        if 'all' in args.strategies:
            strategies_to_run = StrategyRegistry.list_strategies()
        else:
            strategies_to_run = [s for s in args.strategies if s in StrategyRegistry.list_strategies()]

        logger.info(f"即将回测的策略: {', '.join(strategies_to_run)}")

        # 应用策略
        if args.parallel and len(strategies_to_run) > 1:
            # 并行处理策略
            with concurrent.futures.ThreadPoolExecutor(max_workers=min(5, len(strategies_to_run))) as executor:
                futures = []
                for strategy_name in strategies_to_run:
                    strategy_func = StrategyRegistry.get_strategy(strategy_name)
                    if strategy_func:
                        futures.append(executor.submit(
                            backtest.apply_strategy,
                            return_data,
                            strategy_func,
                            strategy_name
                        ))

                # 等待所有任务完成
                for future in concurrent.futures.as_completed(futures):
                    try:
                        future.result()
                    except Exception as e:
                        logger.error(f"并行处理策略时出错: {e}")
        else:
            # 串行处理策略
            for strategy_name in strategies_to_run:
                strategy_func = StrategyRegistry.get_strategy(strategy_name)
                if strategy_func:
                    logger.info(f"开始回测策略: {strategy_name}")
                    backtest.apply_strategy(return_data, strategy_func, strategy_name)

        # 生成报告
        backtest.generate_report(start_date=args.start_date, end_date=end_date)

        # 获取当日匹配股票，并按需导出/发送邮件
        today_matches = None
        latest_date = None
        if args.send_email or args.export_candidates_json:
            today_matches, latest_date = backtest.get_today_matching_stocks(args.strategies)

            if args.export_candidates_json:
                backtest.export_today_matches_json(today_matches, latest_date, args.export_candidates_json)

                # === 新增：调用 /opt/daily_stock_analysis 完成候选增强打分 ===
                enhancer_script = '/opt/daily_stock_analysis/scripts/candidate_enhancer.py'
                if os.path.exists(enhancer_script):
                    try:
                        # 计算输出路径，和输入同一个目录，替换文件名后缀
                        output_json = os.path.splitext(args.export_candidates_json)[0] + '_enhanced.json'
                        output_csv = os.path.splitext(args.export_candidates_json)[0] + '_enhanced.csv'
                        # 使用 conda 调用 daily_stock_analysis 环境执行增强
                        cmd = [
                            '/root/anaconda3/bin/conda', 'run', '-n', 'daily_stock_analysis',
                            'python', enhancer_script,
                            '--input', args.export_candidates_json,
                            '--output', output_json,
                            '--output_csv', output_csv,
                            '--top_n', str(10),
                        ]
                        logger.info('开始执行候选股增强打分: %s', ' '.join(cmd))
                        result = subprocess.run(cmd, check=True, capture_output=True, text=True)
                        if result.returncode == 0:
                            logger.info('候选股增强打分完成: %s', output_json)
                            # 将增强结果路径覆盖，邮件会读取这个增强结果放到顶部
                            args.export_candidates_json = output_json
                        else:
                            logger.error('候选股增强打分失败: retcode=%d', result.returncode)
                    except Exception as e:
                        logger.error('执行候选股增强出错: %s', str(e))
                        logger.error(traceback.format_exc())
                else:
                    logger.warning('候选增强脚本 %s 不存在，跳过增强步骤', enhancer_script)

            if args.send_email:
                # 如果导出了候选股JSON，同时将增强结果放到邮件正文顶部
                backtest.send_strategy_email(today_matches, latest_date, args.export_candidates_json)

    except Exception as e:
        logger.error(f"回测过程中发生错误: {str(e)}")
        logger.error(traceback.format_exc())
    finally:
        backtest.close_pool()


def is_trading_day(date):
    """
    判断指定日期是否为交易日

    Args:
        date (str or datetime): 日期，格式为'YYYY-MM-DD'

    Returns:
        bool: True表示是交易日，False表示非交易日
    """
    # 初始化baostock连接
    lg = bs.login()
    if lg.error_code != '0':
        logger.warning("baostock未正确初始化，无法判断交易日")
        return True  # 出错时默认返回True以保证程序继续运行

    # 格式化日期
    if isinstance(date, datetime):
        date_str = date.strftime('%Y-%m-%d')
    else:
        date_str = str(date)

    # 查询交易日
    rs = bs.query_trade_dates(start_date=date_str, end_date=date_str)

    # 处理查询结果
    if rs.error_code != '0':
        logger.warning(f"查询交易日失败: {rs.error_msg}")
        return True  # 出错时默认返回True

    # 获取查询结果
    data_list = []
    while (rs.error_code == '0') & rs.next():
        data_list.append(rs.get_row_data())

    if not data_list:
        logger.warning(f"未获取到交易日数据: {date_str}")
        return True  # 出错时默认返回True

    # 判断是否为交易日
    is_trading = data_list[0][1] == '1'  # 第二列是is_trading_day字段，'1'表示交易日
    return is_trading

if __name__ == "__main__":
    main()