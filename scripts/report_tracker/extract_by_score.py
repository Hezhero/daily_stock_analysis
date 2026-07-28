#!/opt/daily_stock_analysis/venv/bin/python3
# -*- coding: utf-8 -*-
"""
按评分筛选股票追踪脚本
从 QQ 邮箱读取5月20号之后的股票智能分析报告邮件，
提取评分超过60分的股票，生成追踪Excel。
"""

import imaplib, re, base64, os, sys, glob, shutil, json
from email import header as email_header
from email import message_from_bytes
from datetime import datetime, timedelta
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import baostock as bs

# ── 配置 ──────────────────────────────────────────────────────────────
EMAIL_USER = "851448443@qq.com"
EMAIL_PASS = "ruykqacqbygwbgag"
IMAP_HOST  = "imap.qq.com"
IMAP_PORT  = 993
OUTPUT_DIR = "/opt/daily_stock_analysis/scripts/report_tracker"
os.makedirs(OUTPUT_DIR, exist_ok=True)

CACHE_FILE = os.path.join(OUTPUT_DIR, "price_cache.json")
SCORE_CACHE_FILE = os.path.join(OUTPUT_DIR, "score_cache.json")
CONFIG_FILE = os.path.join(OUTPUT_DIR, "stock_config.json")

# ── 工具函数 ──────────────────────────────────────────────────────────

def stock_suffix(code):
    if code.startswith(('6', '5', '9')):
        return f"sh.{code}"
    return f"sz.{code}"

def decode_mixed_payload(msg):
    for part in msg.walk():
        if part.get_content_type() == 'text/html':
            cte = part.get('Content-Transfer-Encoding', '').lower()
            payload = part.get_payload()
            if cte == 'base64':
                return base64.b64decode(payload.encode('ascii')).decode('utf-8', errors='replace')
            return payload
    return ""

def extract_stocks_with_scores(html):
    """
    从邮件 HTML 中提取个股决策仪表盘中的股票代码和评分
    返回: [(code, score, name), ...] 评分>=60的股票列表
    
    邮件结构:
    ## ⚪ 股票名称 (股票代码)
    > 日期 | 评分: **XX** | 状态
    """
    stocks = []
    
    # 查找"个股决策仪表盘"部分
    dashboard_idx = html.find('个股决策仪表盘')
    if dashboard_idx < 0:
        print(f"    Warning: 未找到'个股决策仪表盘'部分")
        return []
    
    # 只分析仪表盘部分的内容
    dashboard_html = html[dashboard_idx:]
    
    # 查找所有股票区块 - 模式: ## ⚪ 股票名称 (股票代码)
    # 或者: <h2>⚪ 股票名称 (股票代码)</h2>
    stock_patterns = [
        r'<h2[^>]*>[⚪🔴🟢🟡\s]*([^<(]+)\s*\((\d{6})\)</h2>',  # <h2>股票名称 (代码)</h2>
        r'<h2[^>]*>[⚪🔴🟢🟡\s]*([^<(]+)\s*（(\d{6})）',  # 中文括号
    ]
    
    stock_blocks = []
    for pattern in stock_patterns:
        matches = list(re.finditer(pattern, dashboard_html))
        for match in matches:
            name = match.group(1).strip()
            code = match.group(2)
            # 验证股票代码
            if code[0] in '02367':
                start_pos = match.start()
                # 找到下一个h2或h1作为结束位置
                next_h = re.search(r'<h[12][^>]*>', dashboard_html[start_pos+10:])
                end_pos = start_pos + 10 + next_h.start() if next_h else len(dashboard_html)
                block = dashboard_html[start_pos:end_pos]
                stock_blocks.append((name, code, block))
    
    if not stock_blocks:
        print(f"    Warning: 未找到股票区块")
        return []
    
    print(f"    找到 {len(stock_blocks)} 只股票")
    
    # 从每个区块中提取评分
    for name, code, block in stock_blocks:
        # 查找评分模式: 评分: **XX** 或 评分: <strong>XX</strong>
        score_patterns = [
            r'评分[:\s]*(?:<strong>|\*\*)(\d+)(?:</strong>|\*\*)',
            r'评分[:\s]*(\d+)',
        ]
        
        score = None
        for pattern in score_patterns:
            match = re.search(pattern, block)
            if match:
                score = int(match.group(1))
                break
        
        if score is not None:
            print(f"      {name}({code}): 评分={score}")
            if score >= 60:
                stocks.append((code, score, name))
        else:
            print(f"      {name}({code}): 未找到评分")
    
    print(f"    评分>=60的股票: {len(stocks)} 只")
    return stocks

def get_date_from_subject(subject):
    m = re.search(r'(\d{4}-\d{2}-\d{2})', subject)
    return m.group(1) if m else None

def is_trading_day(date_str):
    return datetime.strptime(date_str, '%Y-%m-%d').weekday() < 5

# ── 读取邮件 ─────────────────────────────────────────────────────────────

def read_reports_with_scores():
    """读取邮件并提取带评分的股票数据"""
    conn = imaplib.IMAP4_SSL(IMAP_HOST, IMAP_PORT)
    conn.login(EMAIL_USER, EMAIL_PASS)
    conn.select('"Sent Messages"', readonly=True)
    status, msgs = conn.search(None, 'ALL')
    all_ids = msgs[0].split()

    reports = {}  # {date: [(code, score, name), ...]}
    
    print(f"    总共 {len(all_ids)} 封邮件")
    
    for mid_bytes in reversed(all_ids):  # 从最新的开始
        mid = mid_bytes.decode() if isinstance(mid_bytes, bytes) else str(mid_bytes)
        try:
            # 获取邮件头部
            status, data = conn.fetch(mid, 'BODY.PEEK[HEADER]')
            if status != 'OK':
                continue
                
            raw_header = data[0][1] if isinstance(data[0], tuple) else data[0]
            if isinstance(raw_header, bytes):
                header_str = raw_header.decode('utf-8', errors='replace')
            else:
                header_str = str(raw_header)
            
            msg = message_from_bytes(raw_header if isinstance(raw_header, bytes) else raw_header.encode('utf-8'))
            subj = msg.get('Subject', '')
            
            subj_decoded = ''.join(
                part.decode(enc or 'utf-8', errors='replace') if isinstance(part, bytes) else str(part)
                for part, enc in email_header.decode_header(subj)
            )
            
            if '股票智能分析报告' not in subj_decoded:
                continue
                
            report_date = get_date_from_subject(subj_decoded)
            if not report_date or report_date < '2026-05-20':
                continue
            
            # 如果已经处理过这个日期的邮件，跳过（只保留最后一封）
            if report_date in reports:
                continue

            print(f"\n  处理邮件: {subj_decoded}")
            
            status, data = conn.fetch(mid, 'BODY[]')
            raw_email = data[0][1]
            raw_str = raw_email.decode('utf-8', errors='replace') if isinstance(raw_email, bytes) else raw_email
            m2 = re.match(r'^\d+ \(BODY\[\] \d+\)\r\n', raw_str)
            if m2:
                raw_str = raw_str[m2.end():]
            msg = message_from_bytes(raw_str.encode('utf-8'))
            html = decode_mixed_payload(msg)
            
            stocks = extract_stocks_with_scores(html)
            if stocks:
                reports[report_date] = stocks
                print(f"  [{report_date}] 保存 {len(stocks)} 只股票")
        except Exception as e:
            print(f"    Error processing email {mid}: {e}")
            import traceback
            traceback.print_exc()
            continue
    
    conn.logout()
    return reports

# ── baostock 价格查询 ─────────────────────────────────────────────────

def fetch_prices(codes, start_date, end_date):
    result = {}
    bs.login()
    try:
        for code in codes:
            rs = bs.query_history_k_data_plus(
                stock_suffix(code), 'date,close',
                start_date=start_date, end_date=end_date,
                frequency='d', adjustflag='2')
            while rs.error_code == '0' and rs.next():
                row = rs.get_row_data()
                if row[1]:
                    try:
                        result[(code, row[0])] = round(float(row[1]), 2)
                    except:
                        pass
    finally:
        bs.logout()
    return result

def fetch_names(codes):
    result = {}
    bs.login()
    try:
        for code in codes:
            rs = bs.query_stock_basic(code=stock_suffix(code))
            while rs.error_code == '0' and rs.next():
                row = rs.get_row_data()
                result[code.split('.')[-1]] = row[1]
    finally:
        bs.logout()
    return result

# ── 缓存读写 ─────────────────────────────────────────────────────────────

def load_json(path):
    if os.path.exists(path):
        try:
            with open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    return {}

def save_json(path, data):
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# ── Excel 生成 ─────────────────────────────────────────────────────────

def make_excel(stock_list, price_cache, dates, stock_names, first_rec, scores, out_path):
    """
    生成Excel，包含评分列
    stock_list: [code, ...]
    scores: {code: score}
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "高分股票追踪"

    hdr_fill   = PatternFill("solid", fgColor="1F4E79")
    hdr_font   = Font(color="FFFFFF", bold=True, size=10)
    score_fill = PatternFill("solid", fgColor="FFD700")  # 金色背景用于评分
    alt_fill   = PatternFill("solid", fgColor="EEF4FA")
    green_fill = PatternFill("solid", fgColor="E2EFDA")
    red_fill   = PatternFill("solid", fgColor="FCE4D6")
    thin       = Side(style='thin', color='BBBBBB')
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    center     = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[1].height = 28
    headers = ["股票名称", "股票代码", "评分", "首次入选日期"] + [d for d in dates]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border

    for row, code in enumerate(stock_list, 2):
        ws.row_dimensions[row].height = 20
        
        c1 = ws.cell(row=row, column=1, value=stock_names.get(code, ''))
        c1.alignment = center
        c1.border = border

        c2 = ws.cell(row=row, column=2, value=code)
        c2.alignment = center
        c2.border = border

        c3 = ws.cell(row=row, column=3, value=scores.get(code, ''))
        c3.alignment = center
        c3.border = border
        c3.fill = score_fill  # 评分列用金色背景

        c4 = ws.cell(row=row, column=4, value=first_rec.get(code, ''))
        c4.alignment = center
        c4.border = border

        last_price = None
        for col, date_str in enumerate(dates, 5):
            rec = first_rec.get(code, '9999-99-99')
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = center

            if date_str < rec:
                cell.value = ''
                cell.font = Font(color='CCCCCC')
            else:
                price = price_cache.get(date_str, {}).get(code)
                if price is not None:
                    cell.value = price
                    cell.number_format = '0.00'
                    if last_price is not None:
                        if price > last_price:
                            cell.fill = green_fill
                        elif price < last_price:
                            cell.fill = red_fill
                    last_price = price
                else:
                    cell.value = '—'
                    cell.font = Font(color='BBBBBB')

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 11
    ws.column_dimensions['C'].width = 8   # 评分列
    ws.column_dimensions['D'].width = 12
    for i in range(5, 5 + len(dates)):
        ws.column_dimensions[get_column_letter(i)].width = 12
    ws.freeze_panes = "E2"
    wb.save(out_path)

# ── 主程序 ─────────────────────────────────────────────────────────────

def main():
    t0 = datetime.now()
    print(f"[{t0.strftime('%H:%M:%S')}] === Score-based Stock Tracker ===")

    # 1. 读邮件，提取带评分的股票
    print("[1] Reading emails and extracting scores...")
    reports = read_reports_with_scores()  # {date: [(code, score, name), ...]}
    trading_reports = {d: c for d, c in reports.items() if is_trading_day(d)}
    print(f"\n    Reports found: {sorted(trading_reports.keys())}")

    # 2. 收集所有股票（按首次推荐日期排序）
    first_rec = {}
    stock_scores = {}
    stock_names = {}
    stock_list = []
    seen = set()
    
    for date in sorted(trading_reports.keys()):
        for code, score, name in trading_reports[date]:
            if code not in seen:
                seen.add(code)
                stock_list.append(code)
                first_rec[code] = date
                stock_scores[code] = score
                if name:
                    stock_names[code] = name
    
    print(f"    Total unique stocks (score>=60): {len(stock_list)}")
    if stock_list:
        print(f"    Stocks: {stock_list}")

    # 3. 加载缓存
    price_cache = load_json(CACHE_FILE)
    config = load_json(CONFIG_FILE)
    cached_names = config.get('names', {})
    stock_names.update(cached_names)

    # 4. 确定价格日期范围
    trading_dates = sorted(trading_reports.keys())
    today = datetime.now().strftime('%Y-%m-%d')
    start_d = trading_dates[0] if trading_dates else today
    end_d = today

    # 5. 查询新价格
    dates_to_fetch = [d for d in [start_d, end_d] if d not in price_cache or len(price_cache.get(d, {})) < len(stock_list)]
    if not dates_to_fetch:
        print(f"[2] No new prices needed")
    else:
        print(f"[2] Fetching prices for {dates_to_fetch} ...")
        for d in dates_to_fetch:
            new_prices = fetch_prices(stock_list, d, d)
            if new_prices:
                price_cache[d] = price_cache.get(d, {})
                price_cache[d].update({code: p for (code, _), p in new_prices.items()})
                print(f"    {d}: {len(new_prices)} records")

    # 6. 查询缺失的股票名称
    missing_names = [c for c in stock_list if c not in stock_names]
    if missing_names:
        print(f"[3] Fetching {len(missing_names)} names...")
        new_names = fetch_names(missing_names)
        stock_names.update(new_names)
        print(f"    Got {len(new_names)} names")

    # 7. 保存
    save_json(CACHE_FILE, price_cache)
    save_json(SCORE_CACHE_FILE, stock_scores)
    save_json(CONFIG_FILE, {'names': stock_names, 'first_rec': first_rec, 'scores': stock_scores})
    print(f"    Saved cache and config")

    # 8. 生成 Excel
    exclude_dates = ['2026-04-21', '2026-04-22', '2026-04-23', '2026-04-24', '2026-04-27']
    cached_dates = sorted([d for d in price_cache if is_trading_day(d) and d <= today and d not in exclude_dates],
                         key=lambda x: x)
    
    latest = os.path.join(OUTPUT_DIR, "stock_tracker_by_score.xlsx")
    archive = os.path.join(OUTPUT_DIR, f"stock_tracker_by_score_{datetime.now().strftime('%Y%m%d')}.xlsx")
    
    make_excel(stock_list, price_cache, cached_dates, stock_names, first_rec, stock_scores, latest)
    shutil.copy2(latest, archive)
    
    print(f"[4] Excel: {latest}")
    print(f"    Dates: {cached_dates}")
    print(f"[DONE] {datetime.now().strftime('%H:%M:%S')} ({(datetime.now()-t0).seconds}s)")

if __name__ == "__main__":
    main()
