#!/opt/daily_stock_analysis/venv/bin/python3
# -*- coding: utf-8 -*-
"""
回测报告推荐股票追踪脚本
每天 23:10 执行：从 QQ 邮箱读取04-20之后的回测报告，
提取每个交易日最后一封报告中的10支推荐股票，
用 baostock 查询历史收盘价，更新到 Excel 中。
"""

import imaplib, re, base64, os, sys, glob, shutil, json
from email import header as email_header
from email import message_from_bytes
from datetime import datetime, timedelta
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import baostock as bs

# ── 配置 ──────────────────────────────────────────────────────────────
EMAIL_USER = "851448443@qq.com"
EMAIL_PASS = "ruykqacqbygwbgag"
IMAP_HOST  = "imap.qq.com"
IMAP_PORT  = 993
OUTPUT_DIR = "/opt/daily_stock_analysis/scripts/report_tracker"
os.makedirs(OUTPUT_DIR, exist_ok=True)

CACHE_FILE = os.path.join(OUTPUT_DIR, "price_cache.json")
CONFIG_FILE = os.path.join(OUTPUT_DIR, "stock_config.json")

# ── 工具函数 ──────────────────────────────────────────────────────────

def stock_suffix(code):
    if code.startswith(('6', '5', '9')):
        return f"sh.{code}"
    return f"sz.{code}"

def decode_mixed_payload(msg):
    for part in msg.walk():
        if part.get_content_type() == 'text/html':
            cte = part.get('Content-Transfer-Encoding', '').lower()
            payload = part.get_payload()
            if cte == 'base64':
                return base64.b64decode(payload.encode('ascii')).decode('utf-8', errors='replace')
            return payload
    return ""

def extract_recommended_codes(html):
    """从邮件 HTML 中提取推荐股票代码（限第一个表格10支，排除ETF/基金）"""
    idx = html.find('下个交易日推荐股票')
    if idx < 0:
        return []
    section = html[idx:idx+8000]
    table_end = section.find('</table>')
    if table_end > 0:
        section = section[:table_end]
    text = re.sub(r'<[^>]+>', ' ', section)
    text = re.sub(r'\s+', ' ', text)
    codes = re.findall(r'\b(\d{6})\b', text)
    # A股：0=深圳、2=深圳B、3=创业板、6=沪市、7=沪市新股
    valid = [c for c in codes if c and c[0] in '02367']
    return list(dict.fromkeys(valid))[:10]

def get_date_from_subject(subject):
    m = re.search(r'(\d{4}-\d{2}-\d{2})', subject)
    return m.group(1) if m else None

def is_trading_day(date_str):
    return datetime.strptime(date_str, '%Y-%m-%d').weekday() < 5

# ── 读取邮件 ─────────────────────────────────────────────────────────────

def read_reports():
    conn = imaplib.IMAP4_SSL(IMAP_HOST, IMAP_PORT)
    conn.login(EMAIL_USER, EMAIL_PASS)
    conn.select('"Sent Messages"', readonly=True)
    status, msgs = conn.search(None, 'ALL')
    all_ids = msgs[0].split()

    reports = {}
    for mid_bytes in all_ids:
        mid = mid_bytes.decode() if isinstance(mid_bytes, bytes) else str(mid_bytes)
        try:
            status, data = conn.fetch(mid, '(ENVELOPE)')
            env_bytes = data[0]
            if isinstance(env_bytes, (list, tuple)):
                env_bytes = env_bytes[0]
            env_str = env_bytes.decode('utf-8', errors='replace')
            m = re.search(r'ENVELOPE\s+\(NIL\s+"([^"]*)"', env_str)
            if not m:
                continue
            subj_raw = m.group(1)
            subj = ''.join(
                part.decode(enc or 'utf-8', errors='replace') if isinstance(part, bytes) else str(part)
                for part, enc in email_header.decode_header(subj_raw)
            )
            if '策略5日验证回测报告' not in subj:
                continue
            report_date = get_date_from_subject(subj)
            if not report_date or report_date < '2026-04-20':
                continue

            status, data = conn.fetch(mid, 'BODY[]')
            raw_email = data[0][1]
            raw_str = raw_email.decode('utf-8', errors='replace') if isinstance(raw_email, bytes) else raw_email
            m2 = re.match(r'^\d+ \(BODY\[\] \d+\)\r\n', raw_str)
            if m2:
                raw_str = raw_str[m2.end():]
            msg = message_from_bytes(raw_str.encode('utf-8'))
            html = decode_mixed_payload(msg)
            codes = extract_recommended_codes(html)
            if codes and report_date not in reports:
                reports[report_date] = codes
        except:
            continue
    conn.logout()
    return reports

# ── baostock 价格查询 ─────────────────────────────────────────────────

def fetch_prices(codes, start_date, end_date):
    result = {}
    bs.login()
    try:
        for code in codes:
            rs = bs.query_history_k_data_plus(
                stock_suffix(code), 'date,close',
                start_date=start_date, end_date=end_date,
                frequency='d', adjustflag='2')
            while rs.error_code == '0' and rs.next():
                row = rs.get_row_data()
                if row[1]:
                    try:
                        result[(code, row[0])] = round(float(row[1]), 2)
                    except:
                        pass
    finally:
        bs.logout()
    return result

def fetch_names(codes):
    result = {}
    bs.login()
    try:
        for code in codes:
            rs = bs.query_stock_basic(code=stock_suffix(code))
            while rs.error_code == '0' and rs.next():
                row = rs.get_row_data()
                result[code.split('.')[-1]] = row[1]
    finally:
        bs.logout()
    return result

# ── 缓存读写 ─────────────────────────────────────────────────────────────

def load_json(path):
    if os.path.exists(path):
        try:
            with open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    return {}

def save_json(path, data):
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# ── Excel 生成 ─────────────────────────────────────────────────────────

def make_excel(stock_list, price_cache, dates, stock_names, first_rec, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "推荐股票追踪"

    hdr_fill   = PatternFill("solid", fgColor="1F4E79")
    hdr_font   = Font(color="FFFFFF", bold=True, size=10)
    alt_fill   = PatternFill("solid", fgColor="EEF4FA")
    green_fill = PatternFill("solid", fgColor="E2EFDA")
    red_fill   = PatternFill("solid", fgColor="FCE4D6")
    thin       = Side(style='thin', color='BBBBBB')
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    center     = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[1].height = 28
    headers = ["股票代码", "股票名称"] + [d[5:] for d in dates]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border

    for row, code in enumerate(stock_list, 2):
        ws.row_dimensions[row].height = 20
        c1 = ws.cell(row=row, column=1, value=code)
        c1.alignment = center
        c1.border = border

        c2 = ws.cell(row=row, column=2, value=stock_names.get(code, ''))
        c2.alignment = center
        c2.border = border

        last_price = None
        for col, date_str in enumerate(dates, 3):
            rec = first_rec.get(code, '9999-99-99')
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = center

            if date_str < rec:
                # 该股票尚未被推荐，不显示价格
                cell.value = ''
                cell.font = Font(color='CCCCCC')
                # last_price 保持不变（不更新）
            else:
                price = price_cache.get(date_str, {}).get(code)
                if price is not None:
                    cell.value = price
                    cell.number_format = '0.00'
                    if last_price is not None:
                        if price > last_price:
                            cell.fill = green_fill
                        elif price < last_price:
                            cell.fill = red_fill
                    last_price = price
                else:
                    cell.value = '—'
                    cell.font = Font(color='BBBBBB')
                    # last_price 保持不变

    ws.column_dimensions['A'].width = 11
    ws.column_dimensions['B'].width = 14
    for i in range(3, 3 + len(dates)):
        ws.column_dimensions[get_column_letter(i)].width = 12
    ws.freeze_panes = "C2"
    wb.save(out_path)

# ── 主程序 ─────────────────────────────────────────────────────────────

def main():
    t0 = datetime.now()
    print(f"[{t0.strftime('%H:%M:%S')}] === Report Tracker ===")

    # 1. 读邮件
    print("[1] Reading emails...")
    all_reports = read_reports()  # {date: [codes]}
    trading_reports = {d: c for d, c in all_reports.items() if is_trading_day(d)}
    print(f"    Reports: {sorted(trading_reports.keys())}")

    # 2. 收集股票（已在下面按日期排序时完成）

    # 3. 按首次推荐日期排序（保持邮件中的顺序）
    first_rec = {}
    stock_list = []
    seen = set()
    for date in sorted(trading_reports.keys()):
        for code in trading_reports[date]:
            if code not in seen:
                seen.add(code)
                stock_list.append(code)
                first_rec[code] = date
    print(f"    first_rec sample: {dict(list(first_rec.items())[:3])}")

    # 4. 加载缓存
    price_cache = load_json(CACHE_FILE)
    config = load_json(CONFIG_FILE)
    names = config.get('names', {})

    # 5. 确定价格日期范围
    trading_dates = sorted(trading_reports.keys())
    today = datetime.now().strftime('%Y-%m-%d')
    start_d = trading_dates[0] if trading_dates else today
    end_d = today

    # 6. 查询新价格（只查缺失的日期）
    dates_to_fetch = [d for d in [start_d, end_d] if d not in price_cache or len(price_cache.get(d, {})) < len(stock_list)]
    if not dates_to_fetch:
        print(f"[2] No new prices needed")
    else:
        print(f"[2] Fetching prices for {dates_to_fetch} ...")
        for d in dates_to_fetch:
            new_prices = fetch_prices(stock_list, d, d)
            if new_prices:
                price_cache[d] = price_cache.get(d, {})
                price_cache[d].update({code: p for (code, _), p in new_prices.items()})
                print(f"    {d}: {len(new_prices)} records")

    # 7. 查询股票名称（增量）
    missing_names = [c for c in stock_list if c not in names]
    if missing_names:
        print(f"[3] Fetching {len(missing_names)} names...")
        new_names = fetch_names(missing_names)
        names.update(new_names)
        print(f"    Got {len(new_names)} names")

    # 8. 保存
    save_json(CACHE_FILE, price_cache)
    save_json(CONFIG_FILE, {'names': names, 'first_rec': first_rec})
    print(f"    Saved cache ({len(price_cache)} dates) and config")

    # 9. 生成 Excel
    cached_dates = sorted([d for d in price_cache if is_trading_day(d) and d <= today],
                         key=lambda x: x)
    latest = os.path.join(OUTPUT_DIR, "stock_tracker_latest.xlsx")
    archive = os.path.join(OUTPUT_DIR, f"stock_tracker_{datetime.now().strftime('%Y%m%d')}.xlsx")
    make_excel(stock_list, price_cache, cached_dates, names, first_rec, latest)
    shutil.copy2(latest, archive)
    print(f"[4] Excel: {latest}")
    print(f"    Dates: {cached_dates}")
    print(f"[DONE] {datetime.now().strftime('%H:%M:%S')} ({(datetime.now()-t0).seconds}s)")

if __name__ == "__main__":
    main()